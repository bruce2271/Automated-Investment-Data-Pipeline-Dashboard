"""
=============================================================================
  INVESTING DATA COLLECTOR  v3.9  —  Stage 1
  Sources: FRED (macro) · yfinance (forex/crypto/indices, primary)
           Polygon.io (equity prices, fallback for forex/crypto)
           TWSE (flows) · SEC EDGAR (13F)

  SETUP:
      pip install requests fredapi pandas openpyxl yfinance pandas-datareader
      setx FRED_API_KEY    "your_fred_key"
      setx POLYGON_API_KEY "your_polygon_key"

  SCHEDULE: Daily 08:00 Mon-Fri via Windows Task Scheduler
=============================================================================
"""

import os, sys, time, json, logging, requests
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from fredapi import Fred
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# CONFIG
FRED_API_KEY    = os.environ.get("FRED_API_KEY",    "YOUR_FRED_KEY_HERE")
POLYGON_API_KEY = os.environ.get("POLYGON_API_KEY", "YOUR_POLYGON_KEY_HERE")
WORKBOOK_PATH   = Path(__file__).parent / "master_investing.xlsx"
STATE_FILE      = Path(__file__).parent / "state.json"
FRED_CACHE_FILE = Path(__file__).parent / "fred_cache.json"   # fallback if FRED is down
LOG_FILE        = Path(__file__).parent / "collector.log"
HISTORY_MONTHS  = 24
POLY_PAUSE      = 15

# LOGGING
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# STATE
def load_state():
    try:
        return json.loads(STATE_FILE.read_text()) if STATE_FILE.exists() else {}
    except Exception:
        return {}

def save_state(state):
    try:
        STATE_FILE.write_text(json.dumps(state, indent=2))
    except Exception as e:
        log.warning(f"Could not save state: {e}")

# FRED CACHE  — persists last successful fetch so a FRED outage never wipes US Macro
def load_fred_cache() -> dict:
    """Load cached FRED series. Returns {key: [[date_str, value], ...]}."""
    try:
        if FRED_CACHE_FILE.exists():
            return json.loads(FRED_CACHE_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        log.warning(f"Could not load FRED cache: {e}")
    return {}

def save_fred_cache(fred_data: dict) -> None:
    """
    Persist fred_data to disk after a successful fetch.
    Only non-empty Series are written. Merges with whatever is already cached
    so a partial outage (some series OK, some empty) still updates the good ones.
    """
    try:
        existing = load_fred_cache()
        updated = 0
        for key, series in fred_data.items():
            if series is not None and not series.empty:
                pairs = [
                    [str(idx)[:10], round(float(val), 6)]
                    for idx, val in series.items()
                    if val is not None and str(val) not in ("nan", "None")
                ]
                if pairs:
                    existing[key] = pairs
                    updated += 1
        FRED_CACHE_FILE.write_text(json.dumps(existing, indent=2), encoding="utf-8")
        log.info(f"  FRED cache saved: {updated} series updated → {FRED_CACHE_FILE.name}")
    except Exception as e:
        log.warning(f"Could not save FRED cache: {e}")

def restore_from_fred_cache(fred_data: dict) -> dict:
    """
    For every key whose Series is empty/None, substitute the last cached version.
    Logs exactly which series were restored so you know FRED was down.
    """
    cache = load_fred_cache()
    if not cache:
        return fred_data
    restored = []
    for key in list(fred_data.keys()):
        series = fred_data[key]
        if (series is None or (hasattr(series, "empty") and series.empty)) and key in cache:
            try:
                pairs = cache[key]
                idx  = pd.to_datetime([p[0] for p in pairs])
                vals = [p[1] for p in pairs]
                fred_data[key] = pd.Series(vals, index=idx, dtype=float)
                restored.append(key)
            except Exception as e:
                log.warning(f"  FRED cache restore [{key}]: {e}")
    if restored:
        log.warning(f"  FRED OUTAGE — restored {len(restored)} series from cache: {restored}")
    else:
        log.info("  FRED cache: no restoration needed (all series fetched OK)")
    return fred_data

# STYLES
_HDR  = Font(bold=True, color="FFFFFF", size=11)
_BLUE = Font(color="0000FF")
_GRN  = Font(color="008000")
_BLK  = Font(color="000000")
_RED  = Font(color="CC0000", bold=True)
_DARK = PatternFill("solid", fgColor="1F3864")
_LITE = PatternFill("solid", fgColor="D9E1F2")
_ALT  = PatternFill("solid", fgColor="F2F2F2")
_WHT  = PatternFill("solid", fgColor="FFFFFF")
_YLW  = PatternFill("solid", fgColor="FFFF00")
_SEC  = PatternFill("solid", fgColor="2F5496")
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_BDR  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def _hdr(cell):
    cell.font=_HDR; cell.fill=_DARK; cell.alignment=_CTR; cell.border=_BDR

def _sub(cell):
    cell.font=Font(bold=True,color="1F3864",size=10)
    cell.fill=_LITE; cell.alignment=_CTR; cell.border=_BDR

def _dat(cell, alt=False, yellow=False):
    cell.font=_BLK
    cell.fill=_YLW if yellow else (_ALT if alt else _WHT)
    cell.alignment=_CTR; cell.border=_BDR

def _sec(cell):
    cell.font=Font(bold=True,color="FFFFFF",size=10)
    cell.fill=_SEC; cell.alignment=_LFT; cell.border=_BDR

def _autofit(ws, max_w=32):
    for col in ws.columns:
        w=max((len(str(c.value or "")) for c in col),default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(w+4,max_w)

def _fmt_num(cell, val, is_pct=False, is_price=False, is_large=False):
    """Universal number formatter — 2dp always, no zeros hiding true values."""
    if val is None: return
    if is_large:   cell.number_format="#,##0.00"
    elif is_pct:   cell.number_format="0.00"     # e.g. 3.25 displayed as 3.25
    elif is_price: cell.number_format="#,##0.00"
    else:          cell.number_format="0.0000"    # forex rates need 4dp


# FRED COLLECTOR
class FREDCollector:
    US_SERIES = {
        "fed_funds_rate":    "FEDFUNDS",
        "us_10y_yield":      "DGS10",
        "us_2y_yield":       "DGS2",
        "us_3m_tbill":       "DTB3",
        "t10y2y_spread":     "T10Y2Y",
        "t10y3m_spread":     "T10Y3M",
        "cpi_yoy":           "CPIAUCSL",
        "core_pce_yoy":      "PCEPILFE",
        "ppi_yoy":           "PPIACO",
        "gdp_growth":        "A191RL1Q225SBEA",
        "industrial_prod":   "INDPRO",
        "capacity_util":     "TCU",
        "unemployment":      "UNRATE",
        "nonfarm_payrolls":  "PAYEMS",
        "consumer_sentiment":"UMCSENT",
        "personal_saving":   "PSAVERT",
        "credit_spread_hy":  "BAMLH0A0HYM2",
        "credit_spread_ig":  "BAMLC0A0CM",
        "m2_supply":         "M2SL",
    }
    TW_SERIES = {
        # Only series confirmed to exist on FRED
        "tw_cpi": "TWNPCPIPCPPPT",     # CPI annual % (IMF, annual)
        "tw_gdp": "RGDPNATWA666NRUG",  # Real GDP level (annual, Penn WT)
    }

    def __init__(self, api_key):
        self.fred = Fred(api_key=api_key)

    def _get(self, sid, periods=26):
        try:
            return self.fred.get_series(sid).dropna().tail(periods)
        except Exception as e:
            log.warning(f"FRED {sid}: {e}")
            return pd.Series(dtype=float)

    def _yoy(self, sid):
        try:
            s = self.fred.get_series(sid).dropna()
            return s.pct_change(12).dropna().tail(HISTORY_MONTHS)
        except Exception as e:
            log.warning(f"FRED YoY {sid}: {e}")
            return pd.Series(dtype=float)

    def collect_us(self):
        log.info("Fetching US macro from FRED...")
        d = {}
        for k in ["fed_funds_rate","us_10y_yield","us_2y_yield","us_3m_tbill",
                   "t10y2y_spread","t10y3m_spread","unemployment","capacity_util",
                   "consumer_sentiment","personal_saving","credit_spread_hy","credit_spread_ig"]:
            d[k] = self._get(self.US_SERIES[k])
        for k in ["cpi_yoy","core_pce_yoy","ppi_yoy","industrial_prod","m2_supply"]:
            d[k] = self._yoy(self.US_SERIES[k])
        d["gdp_growth"] = self._get(self.US_SERIES["gdp_growth"], periods=8)
        raw = self._get(self.US_SERIES["nonfarm_payrolls"], periods=26)
        d["nonfarm_payrolls"] = raw.diff().dropna().tail(HISTORY_MONTHS)
        ok = sum(1 for v in d.values() if not v.empty)
        log.info(f"  US FRED: {ok}/{len(d)} series fetched")
        # ── Cache & fallback ──────────────────────────────────────────────────
        # Save whatever we got (even partial) so the cache stays as fresh as possible.
        # Then fill any empty series from the cache in case FRED was partially/fully down.
        save_fred_cache(d)
        d = restore_from_fred_cache(d)
        ok_final = sum(1 for v in d.values() if not v.empty)
        if ok_final > ok:
            log.warning(f"  US FRED final: {ok_final}/{len(d)} series "
                        f"({ok_final - ok} restored from cache)")
        return d

    def collect_tw(self):
        log.info("Fetching Taiwan macro from FRED...")
        d = {}
        for k, sid in self.TW_SERIES.items():
            d[k] = self._get(sid)
        ok = sum(1 for v in d.values() if not v.empty)
        log.info(f"  TW FRED: {ok}/{len(d)} series")
        return d

    def collect_tw_yoy(self):
        """Return YoY % change for TW series that are levels (exports, GDP)."""
        d = self.collect_tw()
        result = {}
        for k, s in d.items():
            if k in ("tw_exports", "tw_gdp") and not s.empty:
                result[k] = s.pct_change(12 if k == "tw_exports" else 1).dropna().tail(HISTORY_MONTHS)
            else:
                result[k] = s
        return result


# POLYGON COLLECTOR
class PolygonCollector:
    BASE = "https://api.polygon.io"

    SIGNALS = {
        # Equities
        "DIA":  "Dow Jones (DIA)",
        "SPY":  "S&P 500 (SPY)",
        "QQQ":  "Nasdaq 100 (QQQ)",
        # Bonds
        "TLT":  "20Y Treasury (TLT)",
        "HYG":  "High Yield Bond (HYG)",
        # Commodities
        "GLD":  "Gold (GLD)",
        "USO":  "Crude Oil (USO)",
        # FX
        "UUP":  "USD Index (UUP)",
    }

    CRYPTO = {
        "X:BTCUSD": "Bitcoin (BTC)",
        "X:ETHUSD": "Ethereum (ETH)",
    }

    # FUTURES removed — merged into SIGNALS (no duplicates)

    FOREX_PAIRS = [
        ("EUR","USD","EUR/USD"),
        ("USD","JPY","USD/JPY"),
        ("USD","TWD","USD/TWD"),
        ("USD","CNY","USD/CNY"),
    ]

    SECTORS = {
        "XLK":"Technology",  "XLF":"Financials",   "XLV":"Healthcare",
        "XLE":"Energy",      "XLI":"Industrials",  "XLY":"Consumer Discret.",
        "XLP":"Consumer Staples","XLU":"Utilities","XLB":"Materials",
        "XLRE":"Real Estate","XLC":"Comm. Services",
    }

    US_STOCKS = {
        "Technology":        [("NVDA","NVIDIA"),        ("AAPL","Apple")],
        "Communication":     [("GOOGL","Alphabet"),     ("META","Meta")],
        "Consumer Discret.": [("AMZN","Amazon"),        ("TSLA","Tesla")],
        "Consumer Staples":  [("WMT","Walmart"),        ("PG","P&G")],
        "Healthcare":        [("UNH","UnitedHealth"),   ("LLY","Eli Lilly")],
        "Financials":        [("JPM","JPMorgan"),       ("BAC","Bank of America")],
        "Industrials":       [("CAT","Caterpillar"),    ("HON","Honeywell")],
        "Energy":            [("XOM","ExxonMobil"),     ("CVX","Chevron")],
        "Materials":         [("LIN","Linde"),          ("SHW","Sherwin-Williams")],
        "Real Estate":       [("PLD","Prologis"),       ("AMT","American Tower")],
        "Utilities":         [("NEE","NextEra Energy"), ("DUK","Duke Energy")],
        "Semiconductors":    [("AVGO","Broadcom"),      ("TSM","TSMC ADR")],
    }

    TW_STOCKS = {
        "AI晶片":   [("2330.TW","台積電 TSMC"),      ("2454.TW","聯發科 MediaTek")],
        "AI伺服器": [("2382.TW","廣達 Quanta"),       ("3231.TW","緯創 Wistron"),
                     ("2317.TW","和碩 Pegatron"),     ("6669.TW","緯穎 Wiwynn")],
        "AI創意":   [("3661.TW","世芯-KY")],
        "AI PC":    [("2357.TW","華碩 ASUS")],
        "散熱":     [("3017.TW","奇鋐 AVC"),          ("3573.TW","雙鴻 Auras"),
                     ("2376.TW","技嘉 Gigabyte")],
        "板卡":     [("2383.TW","台光電 Elite")],
        "電子":     [("2303.TW","聯電 UMC"),          ("3711.TW","日月光 ASE"),
                     ("2308.TW","台達電 Delta")],
        "重電":     [("1513.TW","中興電"),            ("1519.TW","華城電機")],
        "通信網路": [("2412.TW","中華電 CHT"),        ("3045.TW","台灣大 TWM"),
                     ("2881.TW","富邦金")],
        "金融":     [("2882.TW","國泰金"),            ("2891.TW","中信金")],
        "紡織":     [("1476.TW","儒鴻 Eclat"),        ("1477.TW","聚陽 Makalot")],
        "塑化":     [("1301.TW","台塑"),              ("1326.TW","台化"),
                     ("1303.TW","南亞")],
        "水泥":     [("1101.TW","台泥 TCC")],
        "鋼鐵":     [("2002.TW","中鋼 CSC")],
        "汽車":     [("2207.TW","和泰車 Hotai")],
        "食品":     [("1216.TW","統一 Uni-President")],
        "零售":     [("2912.TW","統一超 7-Eleven"),   ("2903.TW","全家 FamilyMart")],
        "數位":     [("2049.TW","富邦媒 momo"),       ("6176.TW","瑞昱 Realtek")],
        "航運":     [("2603.TW","長榮海運 EMC"),      ("2609.TW","陽明海運 YML")],
        "航空":     [("2610.TW","華航 CAL"),          ("2618.TW","長榮航 EVA")],
    }

    def __init__(self, api_key):
        self.api_key = api_key
        self.session = requests.Session()
        self.session.headers.update({"Authorization": f"Bearer {api_key}"})

    def _aggs(self, ticker):
        date_to   = (datetime.today()-timedelta(days=1)).strftime("%Y-%m-%d")
        date_from = (datetime.today()-timedelta(days=400)).strftime("%Y-%m-%d")
        url = f"{self.BASE}/v2/aggs/ticker/{ticker}/range/1/day/{date_from}/{date_to}"
        params = {"adjusted":"true","sort":"asc","limit":365}
        try:
            r = self.session.get(url, params=params, timeout=15)
            if r.status_code == 403:
                log.warning(f"  [{ticker}] 403"); return None
            if r.status_code == 429:
                log.warning(f"  [{ticker}] 429 waiting 60s...")
                time.sleep(60)
                r = self.session.get(url, params=params, timeout=15)
            r.raise_for_status()
            results = r.json().get("results", [])
            if not results: return None
            closes=[b["c"] for b in results]
            highs =[b["h"] for b in results]
            lows  =[b["l"] for b in results]
            dates =[b["t"] for b in results]   # epoch ms
            p=closes[-1]
            hi=max(highs); lo=min(lows)
            p1m=closes[-22] if len(closes)>=22 else None
            p1w=closes[-5]  if len(closes)>=5  else None
            # YTD: find first trading day of current year
            jan1_ts = datetime(datetime.today().year,1,1).timestamp()*1000
            ytd_closes=[closes[j] for j,t in enumerate(dates) if t>=jan1_ts]
            p_ytd=ytd_closes[0] if ytd_closes else closes[0]
            return {
                "last_price":  round(p,2),
                "chg_1w_pct":  round((p/p1w-1)*100,2) if p1w else None,
                "chg_1m_pct":  round((p/p1m-1)*100,2) if p1m else None,
                "chg_ytd_pct": round((p/p_ytd-1)*100,2),
                "high_52w":    round(hi,2),
                "low_52w":     round(lo,2),
                "pct_from_hi": round((p/hi-1)*100,2),
                "as_of":       datetime.today().strftime("%Y-%m-%d"),
            }
        except Exception as e:
            log.warning(f"  [{ticker}] {e}"); return None

    # Hardcoded fundamentals cache — updated quarterly, sourced from public data
    # Format: ticker -> (P/E, P/B, div_yield_pct)
    FUNDAMENTALS_CACHE = {
        "NVDA": (35.2,  35.1, 0.03), "AAPL": (31.5,  45.2, 0.52),
        "MSFT": (32.8,  11.8, 0.81), "GOOGL":(20.9,   6.1, 0.52),
        "META": (23.4,   7.1, 0.40), "AMZN": (34.2,   8.3, None),
        "TSLA": (75.0,  10.5, None), "AVGO": (24.8,  12.4, 1.38),
        "TSM":  (20.1,   6.8, 1.60), "AMD":  (95.0,   3.8, None),
        "JPM":  (12.8,   2.0, 2.20), "BAC":  (12.1,   1.3, 2.40),
        "WMT":  (35.0,   6.8, 0.93), "PG":   (27.5,   7.9, 2.30),
        "UNH":  (18.2,   4.9, 1.80), "LLY":  (55.0,  18.2, 0.65),
        "CAT":  (15.8,   6.4, 1.60), "HON":  (22.1,   7.5, 2.10),
        "XOM":  (13.5,   2.0, 3.50), "CVX":  (14.2,   1.8, 4.20),
        "LIN":  (31.0,   4.8, 1.20), "SHW":  (28.5,   9.3, 0.85),
        "PLD":  (38.5,   3.1, 3.20), "AMT":  (42.0,   8.2, 3.10),
        "NEE":  (19.8,   2.7, 3.40), "DUK":  (18.5,   1.7, 4.20),
    }

    def _fundamentals(self, ticker):
        """Try Polygon first, fall back to hardcoded cache."""
        url = f"{self.BASE}/v3/reference/tickers/{ticker}"
        try:
            r = self.session.get(url, timeout=10)
            r.raise_for_status()
            res = r.json().get("results", {})
            pe  = res.get("pe_ratio")
            pb  = res.get("priceToBook") or res.get("pb_ratio")
            div = res.get("dividend_yield_pct")
            # If Polygon returns all None, use cache
            if pe is None and pb is None and div is None:
                cache = self.FUNDAMENTALS_CACHE.get(ticker, (None,None,None))
                return {"pe_ratio":cache[0],"pb_ratio":cache[1],"div_yield":cache[2]}
            return {"pe_ratio":pe,"pb_ratio":pb,"div_yield":div}
        except Exception:
            cache = self.FUNDAMENTALS_CACHE.get(ticker, (None,None,None))
            return {"pe_ratio":cache[0],"pb_ratio":cache[1],"div_yield":cache[2]}

    def _forex(self, fc, tc):
        date = (datetime.today()-timedelta(days=3)).strftime("%Y-%m-%d")
        url = f"{self.BASE}/v1/open-close/forex/{fc}/{to_c}/{date}"
        try:
            r = self.session.get(url, timeout=10)
            r.raise_for_status()
            d = r.json()
            return {"last_price": d.get("close"), "as_of": d.get("day", date)}
        except Exception as e:
            log.warning(f"  [forex {fc}/{tc}] {e}"); return None

    def _bulk_yf_fetch(self, syms: list) -> dict:
        """One yfinance bulk download. Returns {sym: price_dict}."""
        import yfinance as yf
        today = datetime.today()
        today_str = today.strftime("%Y-%m-%d")
        start = (today - timedelta(days=400)).strftime("%Y-%m-%d")
        result = {}
        try:
            raw = yf.download(syms, start=start, auto_adjust=True, progress=False, threads=True)
            if raw.empty: log.warning("  yfinance bulk: empty"); return {}
            log.info(f"  yfinance bulk: {len(syms)} tickers")
        except Exception as e:
            log.warning(f"  yfinance bulk failed: {e}"); return {}
        single = len(syms) == 1
        for sym in syms:
            try:
                closes = (raw["Close"] if single else raw["Close"][sym]).dropna()
                highs  = (raw["High"]  if single else raw["High"][sym]).dropna()
                lows   = (raw["Low"]   if single else raw["Low"][sym]).dropna()
                if closes.empty: continue
                p = float(closes.iloc[-1])
                hi = float(highs.max()); lo = float(lows.min())
                p1w = float(closes.iloc[-6])  if len(closes)>=6  else None
                p1m = float(closes.iloc[-22]) if len(closes)>=22 else None
                yr = closes[closes.index >= f"{today.year}-01-01"]
                p_ytd = float(yr.iloc[0]) if not yr.empty else float(closes.iloc[0])
                result[sym] = {
                    "last_price":  round(p,2),
                    "chg_1w_pct":  round((p/p1w-1)*100,2) if p1w else None,
                    "chg_1m_pct":  round((p/p1m-1)*100,2) if p1m else None,
                    "chg_ytd_pct": round((p/p_ytd-1)*100,2),
                    "high_52w": round(hi,2), "low_52w": round(lo,2),
                    "pct_from_hi": round((p/hi-1)*100,2),
                    "as_of": today_str,
                }
            except Exception as e:
                log.warning(f"  bulk [{sym}]: {e}")
        log.info(f"  bulk: {len(result)}/{len(syms)} OK")
        return result

    def _collect_map(self, ticker_map):
        rows = []
        items = list(ticker_map.items())
        for i, (sym, name) in enumerate(items):
            log.info(f"  [{i+1}/{len(items)}] {sym}")
            data = self._aggs(sym)
            if data:
                rows.append({"ticker":sym,"name":name,**data})
            if i < len(items)-1:
                time.sleep(POLY_PAUSE)
        log.info(f"  {len(rows)}/{len(items)} OK")
        return pd.DataFrame(rows)

    def collect_signals(self):
        log.info("Fetching market signals (yfinance bulk)...")
        syms = list(self.SIGNALS.keys())
        bulk = self._bulk_yf_fetch(syms)
        rows = []
        for sym, name in self.SIGNALS.items():
            data = bulk.get(sym)
            if data: log.info(f"  {sym} (bulk): {data['last_price']}")
            else:
                log.warning(f"  {sym}: bulk failed → Polygon fallback")
                data = self._aggs(sym)
                if data: time.sleep(POLY_PAUSE)
            if data: rows.append({"ticker": sym, "name": name, **data})
        log.info(f"  Signals: {len(rows)}/{len(syms)} OK")
        return pd.DataFrame(rows)

    def collect_vix(self) -> float | None:
        """Fetch real VIX index value from yfinance (^VIX). Returns latest close or None."""
        try:
            import yfinance as yf
            t    = yf.Ticker("^VIX")
            hist = t.history(period="5d")
            if hist.empty:
                log.warning("  yfinance ^VIX: empty"); return None
            val = float(hist["Close"].dropna().iloc[-1])
            log.info(f"  Real VIX (yfinance ^VIX): {val:.2f}")
            return round(val, 2)
        except Exception as e:
            log.warning(f"  yfinance ^VIX: {e}"); return None

    # yfinance tickers for crypto
    CRYPTO_YF_MAP = {
        "X:BTCUSD": ("Bitcoin (BTC)",   "BTC-USD"),
        "X:ETHUSD": ("Ethereum (ETH)",  "ETH-USD"),
    }

    def _crypto_yf(self, poly_ticker: str, name: str, yf_ticker: str) -> dict | None:
        """Fetch a crypto asset from yfinance. Returns standardised dict or None."""
        try:
            import yfinance as yf
            t    = yf.Ticker(yf_ticker)
            hist = t.history(period="1y")
            if hist.empty:
                log.warning(f"  yfinance crypto [{yf_ticker}]: empty"); return None
            closes = hist["Close"].dropna()
            highs  = hist["High"].dropna()
            lows   = hist["Low"].dropna()
            if closes.empty: return None
            p    = float(closes.iloc[-1])
            hi   = float(highs.max()); lo = float(lows.min())
            p1w  = float(closes.iloc[-6])  if len(closes) >= 6  else None
            p1m  = float(closes.iloc[-22]) if len(closes) >= 22 else None
            yr   = hist[hist.index >= f"{datetime.today().year}-01-01"]["Close"]
            p_ytd = float(yr.iloc[0]) if not yr.empty else float(closes.iloc[0])
            as_of = str(closes.index[-1].date())
            return {
                "ticker":      poly_ticker,
                "name":        name,
                "last_price":  round(p, 2),
                "chg_1w_pct":  round((p/p1w-1)*100, 2) if p1w else None,
                "chg_1m_pct":  round((p/p1m-1)*100, 2) if p1m else None,
                "chg_ytd_pct": round((p/p_ytd-1)*100, 2),
                "high_52w":    round(hi, 2),
                "low_52w":     round(lo, 2),
                "pct_from_hi": round((p/hi-1)*100, 2),
                "as_of":       as_of,
            }
        except Exception as e:
            log.warning(f"  yfinance crypto [{yf_ticker}]: {e}"); return None

    def collect_crypto(self):
        log.info("Fetching crypto (yfinance → Polygon fallback)...")
        rows = []
        items = list(self.CRYPTO_YF_MAP.items())
        for i, (poly_ticker, (name, yf_ticker)) in enumerate(items):
            log.info(f"  [{i+1}/{len(items)}] {poly_ticker}")
            # 1. Try yfinance
            data = self._crypto_yf(poly_ticker, name, yf_ticker)
            if data:
                log.info(f"    {poly_ticker} (yfinance): {data['last_price']}")
                time.sleep(0.3)
            else:
                # 2. Polygon fallback
                raw = self._aggs(poly_ticker)
                if raw:
                    data = {"ticker": poly_ticker, "name": name, **raw}
                    log.info(f"    {poly_ticker} (Polygon): {data['last_price']}")
                    time.sleep(POLY_PAUSE)
            if not data:
                log.warning(f"    {poly_ticker}: all sources failed")
                data = {"ticker":poly_ticker,"name":name,"last_price":None,
                        "chg_1w_pct":None,"chg_1m_pct":None,"chg_ytd_pct":None,
                        "high_52w":None,"low_52w":None,"pct_from_hi":None,"as_of":None}
            rows.append(data)
        return pd.DataFrame(rows)

    # collect_futures() removed — data merged into collect_signals()

    def collect_sectors(self):
        log.info("Fetching sector ETFs (yfinance bulk)...")
        syms = list(self.SECTORS.keys())
        bulk = self._bulk_yf_fetch(syms)
        rows = []
        for sym, name in self.SECTORS.items():
            data = bulk.get(sym)
            if data: log.info(f"  {sym} (bulk): {data['last_price']}")
            else:
                log.warning(f"  {sym}: bulk failed → Polygon fallback")
                data = self._aggs(sym)
                if data: time.sleep(POLY_PAUSE)
            if data: rows.append({"ticker": sym, "name": name, **data})
        log.info(f"  Sectors: {len(rows)}/{len(syms)} OK")
        return pd.DataFrame(rows)

    # yfinance tickers for forex pairs (Yahoo Finance format)
    FOREX_YF_MAP = {
        "EUR/USD": "EURUSD=X",
        "USD/JPY": "JPY=X",
        "USD/TWD": "TWD=X",
        "USD/CNY": "CNY=X",
    }
    # Polygon C: tickers as fallback
    FOREX_POLY_MAP = {
        "EUR/USD": "C:EURUSD",
        "USD/JPY": "C:USDJPY",
        "USD/TWD": "C:USDTWD",
        "USD/CNY": "C:USDCNY",
    }

    def _forex_yf(self, label: str, yf_ticker: str) -> dict | None:
        """Fetch a forex pair from yfinance. Returns standardised dict or None."""
        try:
            import yfinance as yf
            t    = yf.Ticker(yf_ticker)
            hist = t.history(period="1y")
            if hist.empty:
                log.warning(f"  yfinance forex [{yf_ticker}]: empty"); return None
            closes = hist["Close"].dropna()
            if closes.empty: return None
            p    = float(closes.iloc[-1])
            p1w  = float(closes.iloc[-6])  if len(closes) >= 6  else None
            p1m  = float(closes.iloc[-22]) if len(closes) >= 22 else None
            yr   = hist[hist.index >= f"{datetime.today().year}-01-01"]["Close"]
            p_ytd = float(yr.iloc[0]) if not yr.empty else float(closes.iloc[0])
            as_of = str(closes.index[-1].date())
            return {
                "pair":       label,
                "last_price": round(p, 4),
                "chg_1w_pct": round((p/p1w-1)*100, 2) if p1w else None,
                "chg_1m_pct": round((p/p1m-1)*100, 2) if p1m else None,
                "chg_ytd_pct":round((p/p_ytd-1)*100, 2),
                "as_of":      as_of,
            }
        except Exception as e:
            log.warning(f"  yfinance forex [{yf_ticker}]: {e}"); return None

    def collect_forex(self):
        log.info("Fetching forex pairs (yfinance → Polygon fallback)...")
        rows = []
        for i, (label, yf_ticker) in enumerate(self.FOREX_YF_MAP.items()):
            log.info(f"  [{i+1}/{len(self.FOREX_YF_MAP)}] {label}")
            # 1. Try yfinance
            data = self._forex_yf(label, yf_ticker)
            if data:
                log.info(f"    {label} (yfinance): {data['last_price']}")
                time.sleep(0.3)
            else:
                # 2. Polygon fallback
                poly_ticker = self.FOREX_POLY_MAP.get(label)
                if poly_ticker:
                    raw = self._aggs(poly_ticker)
                    if raw and raw.get("last_price"):
                        data = {
                            "pair":       label,
                            "last_price": round(raw["last_price"], 4),
                            "chg_1w_pct": raw.get("chg_1w_pct"),
                            "chg_1m_pct": raw.get("chg_1m_pct"),
                            "chg_ytd_pct":raw.get("chg_ytd_pct"),
                            "as_of":      raw.get("as_of"),
                        }
                        log.info(f"    {label} (Polygon): {data['last_price']}")
                        time.sleep(POLY_PAUSE)
            if not data:
                log.warning(f"    {label}: all sources failed")
                data = {"pair":label,"last_price":None,"chg_1w_pct":None,
                        "chg_1m_pct":None,"chg_ytd_pct":None,"as_of":None}
            rows.append(data)
        return pd.DataFrame(rows)

    def _yf_aggs(self, sym: str) -> dict | None:
        """yfinance price data — same shape as _aggs(). Used as fallback for US/TW stocks."""
        try:
            import yfinance as yf
            t = yf.Ticker(sym); hist = t.history(period="1y")
            if hist.empty: return None
            closes = hist["Close"].dropna()
            highs = hist["High"].dropna(); lows = hist["Low"].dropna()
            if closes.empty: return None
            p = float(closes.iloc[-1])
            hi = float(highs.max()); lo = float(lows.min())
            p1w = float(closes.iloc[-6])  if len(closes)>=6  else None
            p1m = float(closes.iloc[-22]) if len(closes)>=22 else None
            yr = hist[hist.index >= f"{datetime.today().year}-01-01"]["Close"]
            p_ytd = float(yr.iloc[0]) if not yr.empty else float(closes.iloc[0])
            return {
                "last_price":  round(p,2),
                "chg_1w_pct":  round((p/p1w-1)*100,2) if p1w else None,
                "chg_1m_pct":  round((p/p1m-1)*100,2) if p1m else None,
                "chg_ytd_pct": round((p/p_ytd-1)*100,2),
                "high_52w": round(hi,2), "low_52w": round(lo,2),
                "pct_from_hi": round((p/hi-1)*100,2),
                "as_of": str(closes.index[-1].date()),
            }
        except Exception as e:
            log.warning(f"  yfinance [{sym}]: {e}"); return None

    def _collect_stocks_by_sector(self, sector_map):
        rows = []
        all_tickers = [(sec,sym,name) for sec,lst in sector_map.items() for sym,name in lst]
        for i, (sector, sym, name) in enumerate(all_tickers):
            log.info(f"  [{i+1}/{len(all_tickers)}] {sym} ({sector})")
            is_tw = ".TW" in sym
            data  = None
            if is_tw:
                data = self._yf_aggs(sym)
                if data: log.info(f"    {sym} (yfinance): {data['last_price']}")
                else:    log.warning(f"    {sym}: yfinance failed")
            else:
                data = self._aggs(sym)
                if data:
                    log.info(f"    {sym} (Polygon): {data['last_price']}")
                    time.sleep(POLY_PAUSE)
                else:
                    log.warning(f"    {sym}: Polygon failed → yfinance fallback")
                    data = self._yf_aggs(sym)
                    if data: log.info(f"    {sym} (yfinance): {data['last_price']}")
                    time.sleep(0.5)
            if data is None:
                if i < len(all_tickers)-1 and not is_tw: time.sleep(POLY_PAUSE)
                continue
            fund = self._fundamentals(sym) if not is_tw else {"pe_ratio":None,"pb_ratio":None,"div_yield":None}
            rows.append({"sector":sector,"ticker":sym,"name":name,**data,**fund})
            if i < len(all_tickers)-1 and not is_tw: time.sleep(POLY_PAUSE)
        log.info(f"  {len(rows)}/{len(all_tickers)} stocks OK")
        return pd.DataFrame(rows)

    def collect_us_stocks(self):
        log.info("Fetching US stocks (yfinance bulk)...")
        all_entries = [(sec,sym,name) for sec,lst in self.US_STOCKS.items() for sym,name in lst]
        syms = [sym for _,sym,_ in all_entries]
        bulk = self._bulk_yf_fetch(syms)
        rows = []
        for sector, sym, name in all_entries:
            data = bulk.get(sym)
            if data: log.info(f"  {sym} (bulk): {data['last_price']}")
            else:
                log.warning(f"  {sym}: bulk failed → Polygon fallback")
                data = self._aggs(sym)
                if data: log.info(f"  {sym} (Polygon): {data['last_price']}"); time.sleep(POLY_PAUSE)
            if not data: log.warning(f"  {sym}: all sources failed"); continue
            fund = self._fundamentals(sym)
            rows.append({"sector":sector,"ticker":sym,"name":name,**data,**fund})
        log.info(f"  US stocks: {len(rows)}/{len(syms)} OK")
        return pd.DataFrame(rows)

    def collect_tw_stocks(self):
        log.info("Fetching Taiwan stocks (yfinance bulk)...")
        all_entries = [(sec,sym,name) for sec,lst in self.TW_STOCKS.items() for sym,name in lst]
        syms = [sym for _,sym,_ in all_entries]
        bulk = self._bulk_yf_fetch(syms)
        rows = []
        for sector, sym, name in all_entries:
            data = bulk.get(sym)
            if data: log.info(f"  {sym} (bulk): {data['last_price']}")
            else: log.warning(f"  {sym}: yfinance bulk failed"); continue
            rows.append({"sector":sector,"ticker":sym,"name":name,
                         "pe_ratio":None,"pb_ratio":None,"div_yield":None,**data})
        log.info(f"  TW stocks: {len(rows)}/{len(syms)} OK")
        return pd.DataFrame(rows)


# STOOQ INDEX COLLECTOR — uses pandas_datareader which handles Stooq session correctly
class GlobalIndexCollector:
    """
    Fetches global index values from multiple free sources.
    Priority: yfinance → Stooq (pandas_datareader) → Polygon ETF proxy
    No API key required for yfinance or Stooq.
    Requires: pip install yfinance pandas-datareader
    """

    # (ticker, display_name, polygon_etf_fallback)
    INDICES = [
        # US
        ("^GSPC",  "S&P 500",        "SPY"),
        ("^IXIC",  "Nasdaq Composite","QQQ"),
        ("^DJI",   "Dow Jones",       "DIA"),
        ("^RUT",   "Russell 2000",    "IWM"),
        ("^VIX",   "VIX",             "VIXY"),
        # Taiwan
        ("^TWII",  "TAIEX",           None),
        # Asia
        ("^N225",  "Nikkei 225",      None),
        ("^HSI",   "Hang Seng",       None),
        ("^KS11",  "KOSPI",           None),
        ("000001.SS","Shanghai Comp.", None),
        # Europe
        ("^FTSE",  "FTSE 100",        None),
        ("^GDAXI", "DAX",             None),
        ("^FCHI",  "CAC 40",          None),
    ]

    def _fetch_yfinance(self, ticker: str) -> dict | None:
        """Try yfinance with browser-like session to avoid blocks."""
        try:
            import yfinance as yf
            # Let yfinance manage its own session (uses curl_cffi if available)
            # Do NOT pass a custom requests.Session — Yahoo now blocks it
            t    = yf.Ticker(ticker)
            hist = t.history(period="1y")
            if hist.empty:
                log.warning(f"  yfinance [{ticker}]: empty history")
                return None
            closes = hist["Close"].dropna()
            if closes.empty:
                return None
            p   = float(closes.iloc[-1])
            p1w = float(closes.iloc[-6])  if len(closes) >= 6  else None
            p1m = float(closes.iloc[-22]) if len(closes) >= 22 else None
            # YTD: first close of current year
            yr_data = hist[hist.index >= f"{datetime.today().year}-01-01"]
            p_ytd   = float(yr_data["Close"].iloc[0]) if not yr_data.empty else float(closes.iloc[0])
            as_of   = str(closes.index[-1].date())
            return {
                "last_price":  round(p, 2),
                "chg_1w_pct":  round((p/p1w-1)*100, 2) if p1w else None,
                "chg_1m_pct":  round((p/p1m-1)*100, 2) if p1m else None,
                "chg_ytd_pct": round((p/p_ytd-1)*100, 2),
                "as_of":       as_of,
            }
        except ImportError:
            log.warning("  yfinance not installed — run: pip install --upgrade yfinance")
            return None
        except Exception as e:
            log.warning(f"  yfinance [{ticker}]: {e}")
            return None

    def _fetch_stooq(self, ticker: str) -> dict | None:
        """Stooq via pandas_datareader — fallback for indices."""
        try:
            from pandas_datareader.stooq import StooqDailyReader
            end   = datetime.today()
            start = end - timedelta(days=400)
            df    = StooqDailyReader(ticker, start=start, end=end).read()
            if df is None or df.empty:
                return None
            df     = df.sort_index()
            closes = df["Close"].dropna()
            dates  = [str(d.date()) for d in closes.index]
            if closes.empty:
                return None
            p    = float(closes.iloc[-1])
            p1w  = float(closes.iloc[-6])  if len(closes) >= 6  else None
            p1m  = float(closes.iloc[-23]) if len(closes) >= 23 else None
            yr   = str(datetime.today().year)
            ytd  = [float(closes.iloc[i]) for i,d in enumerate(dates) if d.startswith(yr)]
            p_ytd = ytd[0] if ytd else float(closes.iloc[0])
            return {
                "last_price":  round(p, 2),
                "chg_1w_pct":  round((p/p1w-1)*100, 2) if p1w else None,
                "chg_1m_pct":  round((p/p1m-1)*100, 2) if p1m else None,
                "chg_ytd_pct": round((p/p_ytd-1)*100, 2),
                "as_of":       dates[-1],
            }
        except Exception as e:
            log.warning(f"  Stooq [{ticker}]: {e}")
            return None

    def _fetch_twse_taiex(self) -> dict | None:
        """TWSE direct API for TAIEX — free, no key needed."""
        try:
            url = "https://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&type=MS"
            r = requests.get(url, timeout=15, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Referer":    "https://www.twse.com.tw/zh/page/trading/exchange/MI_INDEX.html",
                "Accept":     "application/json, text/javascript, */*",
            })
            r.raise_for_status()
            data = r.json()
            # Find the closing index row
            for row in data.get("data", []):
                if len(row) >= 2 and "加權" in str(row[0]):
                    try:
                        price = float(str(row[1]).replace(",", ""))
                        return {"last_price": price, "chg_1w_pct": None,
                                "chg_1m_pct": None, "chg_ytd_pct": None,
                                "as_of": str(datetime.today().date())}
                    except: pass
            return None
        except Exception as e:
            log.warning(f"  TWSE TAIEX: {e}")
            return None

    def collect(self, poly_collector=None) -> pd.DataFrame:
        log.info("Fetching global indices (yfinance → Stooq → Polygon ETF)...")
        rows = []
        for ticker, name, etf_fallback in self.INDICES:
            data = None
            # 1. Special case: TAIEX from TWSE direct
            if ticker == "^TWII":
                data = self._fetch_twse_taiex()
                if data:
                    log.info(f"  TAIEX (TWSE): {data['last_price']}")
                    time.sleep(1)

            # 2. Try yfinance
            if not data:
                data = self._fetch_yfinance(ticker)
                if data:
                    log.info(f"  {ticker} (yfinance): {data['last_price']}")
                    time.sleep(0.5)

            # 3. Try Stooq
            if not data:
                data = self._fetch_stooq(ticker)
                if data:
                    log.info(f"  {ticker} (Stooq): {data['last_price']}")
                    time.sleep(1)

            # 4. Polygon ETF proxy
            if not data and etf_fallback and poly_collector:
                try:
                    data = poly_collector._aggs(etf_fallback)
                    if data:
                        log.info(f"  {ticker} (Polygon ETF {etf_fallback}): {data['last_price']}")
                        time.sleep(15)
                except Exception as e:
                    log.warning(f"  Polygon ETF {etf_fallback}: {e}")

            if not data:
                log.warning(f"  {ticker}: all sources failed")
                data = {"last_price": None, "chg_1w_pct": None,
                        "chg_1m_pct": None, "chg_ytd_pct": None, "as_of": None}

            rows.append({"ticker": ticker, "name": name, **data})

        ok = sum(1 for r in rows if r.get("last_price"))
        log.info(f"  Global indices: {ok}/{len(rows)} OK")
        return pd.DataFrame(rows)


class TWStockCollector:
    """
    Fetches Taiwan stock prices, P/E, P/B and dividend yield
    directly from TWSE OpenAPI — free, no key, one bulk call covers all stocks.

    Endpoints:
      STOCK_DAY_ALL  : today's OHLC + volume for every listed stock
      BWIBBU_ALL     : P/E, P/B, dividend yield for every listed stock (本益比/殖利率/股價淨值比)
    Both return the full market in a single request — no per-ticker calls needed.
    """
    PRICE_URL = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
    VALUATION_URL = "https://www.twse.com.tw/exchangeReport/BWIBBU_ALL?response=open_data"
    HDR = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}

    # Your watchlist — maps stock code to (sector, display name)
    WATCHLIST = {
        "2330": ("AI晶片",   "台積電 TSMC"),
        "2454": ("AI晶片",   "聯發科 MediaTek"),
        "2382": ("AI伺服器", "廣達 Quanta"),
        "3231": ("AI伺服器", "緯創 Wistron"),
        "2317": ("AI伺服器", "和碩 Pegatron"),
        "6669": ("AI伺服器", "緯穎 Wiwynn"),
        "3661": ("AI創意",   "世芯-KY"),
        "2357": ("AI PC",    "華碩 ASUS"),
        "3017": ("散熱",     "奇鋐 AVC"),
        "3573": ("散熱",     "雙鴻 Auras"),
        "2376": ("散熱",     "技嘉 Gigabyte"),
        "2383": ("板卡",     "台光電 Elite"),
        "2303": ("電子",     "聯電 UMC"),
        "3711": ("電子",     "日月光 ASE"),
        "2308": ("電子",     "台達電 Delta"),
        "1513": ("重電",     "中興電"),
        "1519": ("重電",     "華城電機"),
        "2412": ("通信網路", "中華電 CHT"),
        "3045": ("通信網路", "台灣大 TWM"),
        "2881": ("通信網路", "富邦金"),
        "2882": ("金融",     "國泰金"),
        "2891": ("金融",     "中信金"),
        "1476": ("紡織",     "儒鴻 Eclat"),
        "1477": ("紡織",     "聚陽 Makalot"),
        "1301": ("塑化",     "台塑"),
        "1326": ("塑化",     "台化"),
        "1303": ("塑化",     "南亞"),
        "1101": ("水泥",     "台泥 TCC"),
        "2002": ("鋼鐵",     "中鋼 CSC"),
        "2207": ("汽車",     "和泰車 Hotai"),
        "1216": ("食品",     "統一 Uni-President"),
        "2912": ("零售",     "統一超 7-Eleven"),
        "2903": ("零售",     "全家 FamilyMart"),
        "2049": ("數位",     "富邦媒 momo"),
        "6176": ("數位",     "瑞昱 Realtek"),
        "2603": ("航運",     "長榮海運 EMC"),
        "2609": ("航運",     "陽明海運 YML"),
        "2610": ("航空",     "華航 CAL"),
        "2618": ("航空",     "長榮航 EVA"),
    }

    def _fetch(self, url: str) -> list:
        try:
            r = requests.get(url, headers=self.HDR, timeout=20)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            log.warning(f"  TWSE fetch failed {url}: {e}")
            return []

    def collect(self) -> pd.DataFrame:
        log.info("Fetching Taiwan stock prices from TWSE OpenAPI...")

        # ── Price data (single bulk call) ──
        price_data = self._fetch(self.PRICE_URL)
        price_map = {}
        for row in price_data:
            code = str(row.get("Code","")).strip()
            if code in self.WATCHLIST:
                def _f(v):
                    try: return float(str(v).replace(",",""))
                    except: return None
                price_map[code] = {
                    "close":  _f(row.get("ClosingPrice")),
                    "open":   _f(row.get("OpeningPrice")),
                    "high":   _f(row.get("HighestPrice")),
                    "low":    _f(row.get("LowestPrice")),
                    "change": _f(row.get("Change")),
                    "volume": _f(row.get("TradeVolume")),
                }
        log.info(f"  Price: {len(price_map)}/{len(self.WATCHLIST)} stocks found")

        # ── Valuation: BWIBBU_ALL CSV ──
        # Columns: 日期[0], 股票代號[1], 股票名稱[2], 本益比[3], 殖利率(%)[4], 股價淨值比[5]
        # CRITICAL: utf-8-sig strips the BOM byte that breaks column reading
        time.sleep(2)
        val_map = {}
        try:
            rv = requests.get(self.VALUATION_URL, timeout=20, headers=self.HDR)
            rv.raise_for_status()
            rv.encoding = "utf-8-sig"   # strips BOM if present
            import csv as _csv, io as _io
            def _fv(v):
                s = str(v).strip() if v else ""
                if s in ("", "--", "N/A", "-"): return None
                try: return round(float(s.replace(",","")), 2)
                except: return None
            reader = _csv.reader(_io.StringIO(rv.text))
            header = next(reader, None)
            log.info(f"  BWIBBU_ALL header: {header}")
            for parts in reader:
                if len(parts) < 6: continue
                code = parts[1].strip()
                if code in self.WATCHLIST:
                    val_map[code] = {
                        "pe_ratio":  _fv(parts[3]),
                        "div_yield": _fv(parts[4]),
                        "pb_ratio":  _fv(parts[5]),
                    }
            log.info(f"  BWIBBU_ALL: {len(val_map)}/{len(self.WATCHLIST)} matched")
        except Exception as e:
            log.warning(f"  BWIBBU_ALL failed: {e}")

        # ── Build output rows ──
        today = datetime.today().strftime("%Y-%m-%d")
        rows = []
        for code, (sector, name) in self.WATCHLIST.items():
            p = price_map.get(code, {})
            v = val_map.get(code, {})
            close = p.get("close")
            chg   = p.get("change")
            rows.append({
                "sector":    sector,
                "ticker":    code,
                "name":      name,
                "last_price":close,
                "change":    chg,
                "chg_pct":   round(chg/close*100, 2) if close and chg else None,
                "open":      p.get("open"),
                "high":      p.get("high"),
                "low":       p.get("low"),
                "volume_k":  round(p["volume"]/1000,0) if p.get("volume") else None,
                "pe_ratio":  v.get("pe_ratio"),
                "pb_ratio":  v.get("pb_ratio"),
                "div_yield": v.get("div_yield"),
                "as_of":     today,
            })

        df = pd.DataFrame(rows)
        got = df["last_price"].notna().sum()
        log.info(f"  TW stocks: {got}/{len(rows)} with price data")
        return df

    def collect_taiex(self) -> dict:
        """Fetch TAIEX index value from MI_INDEX endpoint."""
        url = "https://openapi.twse.com.tw/v1/exchangeReport/MI_INDEX"
        data = self._fetch(url)
        # MI_INDEX returns list; find the overall index row (指數 = TAIEX)
        for row in data:
            name = str(row.get("指數名稱","") or row.get("Index",""))
            if "加權" in name or "TAIEX" in name.upper() or "發行量" in name:
                def _f(v):
                    try: return round(float(str(v).replace(",","")),2)
                    except: return None
                return {
                    "name":    name,
                    "close":   _f(row.get("收盤指數") or row.get("Close")),
                    "change":  _f(row.get("漲跌點數") or row.get("Change")),
                    "chg_pct": _f(row.get("漲跌百分比") or row.get("ChangePct")),
                    "as_of":   datetime.today().strftime("%Y-%m-%d"),
                }
        return {}


# TWSE FLOWS
class TWInstitutionalCollector:
    URL = "https://www.twse.com.tw/fund/BFI82U?response=json&dayDate=&type=day"

    def collect(self):
        log.info("Fetching TWSE institutional flows...")
        try:
            r = requests.get(self.URL, timeout=15, headers={"User-Agent":"Mozilla/5.0"})
            r.raise_for_status()
            data = r.json()
            label_map = {
                "自營商(自行買賣)":"Dealer (own)",
                "自營商(避險)":"Dealer (hedge)",
                "投信":"Investment Trust",
                "外資及陸資(不含外資自營商)":"Foreign Investors",
                "合計":"Total",
            }
            result = {"as_of": data.get("date",""), "rows":[]}
            for row in data.get("data",[]):
                if len(row) < 4: continue
                try:
                    name = row[0].strip()
                    buy  = float(str(row[1]).replace(",",""))
                    sell = float(str(row[2]).replace(",",""))
                    net  = float(str(row[3]).replace(",",""))
                    result["rows"].append({
                        "entity":  label_map.get(name, name),
                        "buy_bn":  round(buy/1e8,2),
                        "sell_bn": round(sell/1e8,2),
                        "net_bn":  round(net/1e8,2),
                    })
                except Exception:
                    continue
            log.info(f"  TW flows: {len(result['rows'])} rows")
            return result
        except Exception as e:
            log.warning(f"TWSE flows: {e}")
            return {"as_of":"","rows":[]}

    def collect_stock_flows(self, stock_ids: list) -> list:
        """
        Fetch per-stock 三大法人 flows using TWSE T86 bulk endpoint.
        ONE call returns ALL stocks.

        T86 columns: [0]代號 [1]名稱 [2]外資買 [3]外資賣 [4]外資淨
                     [5]投信買 [6]投信賣 [7]投信淨
                     [8]自營買 [9]自營賣 [10]自營淨 [11]三大合計
        """
        log.info("Fetching per-stock institutional flows via T86 bulk...")
        today = datetime.today()
        # Start from yesterday (offset=1): at 6 AM, today's T86 not published yet.
        # TWSE publishes T86 after ~18:00 Taipei time, so yesterday is always complete.
        offset = 1
        while True:
            candidate = today - timedelta(days=offset)
            if candidate.weekday() < 5: break
            offset += 1
        date_str = candidate.strftime("%Y%m%d")
        log.info(f"  T86 date: {date_str} ({candidate.strftime('%A')})")

        def _n(val):
            s = str(val).replace(",","").replace("+","").strip()
            if s in ("","--","X","N/A"): return None
            try: return round(float(s), 0)
            except: return None

        watch = {sid.replace(".TW","") for sid in stock_ids}
        url = (f"https://www.twse.com.tw/fund/T86"
               f"?response=json&date={date_str}&selectType=ALLBUT0999")
        try:
            r = requests.get(url, timeout=20, headers={"User-Agent":"Mozilla/5.0"})
            r.raise_for_status()
            all_rows = r.json().get("data", [])
            if len(all_rows) == 0:
                log.warning("  T86 returned 0 rows — TWSE publishes after-market data after 15:00 Taipei time. Re-run after market close for institutional flows.")
            else:
                log.info(f"  T86 returned {len(all_rows)} rows")
            row_map = {}
            for row in all_rows:
                if len(row) < 12: continue
                code = str(row[0]).strip()
                if code in watch:
                    row_map[code] = {
                        "stock":       code+".TW",
                        "foreign_net": _n(row[4]),
                        "trust_net":   _n(row[7]),
                        "dealer_net":  _n(row[10]),
                        "total_net":   _n(row[11]),
                        "as_of":       date_str,
                    }
            results = []
            for sid in stock_ids:
                code = sid.replace(".TW","")
                if code in row_map:
                    d = row_map[code]
                    log.info(f"  {code}: F={d['foreign_net']} T={d['trust_net']} Tot={d['total_net']}")
                    results.append(d)
                else:
                    results.append({"stock":sid,"foreign_net":None,"trust_net":None,
                                    "dealer_net":None,"total_net":None,"as_of":date_str})
            found = sum(1 for r in results if r["total_net"] is not None)
            log.info(f"  T86: {found}/{len(stock_ids)} matched")
            return results
        except Exception as e:
            log.warning(f"  T86 failed: {e}")
            return [{"stock":sid,"foreign_net":None,"trust_net":None,
                     "dealer_net":None,"total_net":None,"as_of":date_str}
                    for sid in stock_ids]



# SEC 13F COLLECTOR
class SEC13FCollector:
    """
    Pulls 13F institutional holdings from SEC EDGAR (fully public, no key needed).
    13F is quarterly, filed within 45 days of quarter-end — data is always delayed.
    We track 5 major funds and filter holdings to your US stock watchlist only.

    Funds tracked:
      Berkshire Hathaway : CIK 0001067983
      BlackRock          : CIK 0001364742
      Vanguard           : CIK 0000102909
      State Street       : CIK 0000093751
      Bridgewater        : CIK 0001350694
    """
    EDGAR = "https://data.sec.gov"
    FUNDS = {
        "Berkshire Hathaway": "0001067983",
        "BlackRock":          "0001364742",
        "Vanguard":           "0000102909",
        "State Street":       "0000093751",
        "Bridgewater":        "0001350694",
    }
    WATCH = {
        "NVDA","AAPL","MSFT","GOOGL","META","AMZN","TSLA",
        "AVGO","TSM","AMD","JPM","BAC","WMT","PG",
        "UNH","LLY","XOM","CVX","CAT","HON","NEE","DUK","PLD","AMT","LIN","SHW",
    }

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "InvestingWorksheet contact@personal.com",
            "Accept-Encoding": "gzip, deflate",
        })

    def _latest_13f(self, cik: str):
        """
        Fetch most recent 13F-HR via EDGAR submissions JSON API.
        EDGAR requires a User-Agent header or returns 403.
        """
        cik_int    = str(int(cik))
        cik_padded = cik_int.zfill(10)
        url = f"{self.EDGAR}/submissions/CIK{cik_padded}.json"
        hdrs = {"User-Agent": "investing-collector/1.0 contact@example.com",
                "Accept": "application/json"}
        try:
            r = self.session.get(url, timeout=15, headers=hdrs)
            r.raise_for_status()
            data    = r.json()
            filings = data.get("filings", {}).get("recent", {})
            forms   = filings.get("form", [])
            accnums = filings.get("accessionNumber", [])  # e.g. "0001193125-26-054580"
            dates   = filings.get("filingDate", [])
            primary_docs = filings.get("primaryDocument", [])
            for form, acc, date, primary in zip(forms, accnums, dates, primary_docs or [""]*len(forms)):
                # Match 13F-HR (not 13F-HR/A amendments)
                if str(form).upper().strip() == "13F-HR":
                    result = {
                        "accession_dashed": acc,
                        "accession_clean":  acc.replace("-",""),
                        "cik_int":          cik_int,
                        "filed":            date,
                        "primary_doc":      str(primary) if primary else "",
                    }
                    log.info(f"    Found 13F-HR: {acc} filed {date}")
                    return result
            log.warning(f"  EDGAR: no 13F-HR found for CIK {cik_int}")
            return None
        except Exception as e:
            log.warning(f"  EDGAR submissions CIK{cik_padded}: {e}"); return None

    def _holdings(self, filing: dict) -> list:
        """
        Download 13F info-table XML using approach from SEC developer docs:
        1. Try informationtable.xml (most common standard name)
        2. Try the primaryDocument field from submissions JSON
        3. Fall back to parsing index HTML for any non-primary XML
        """
        cik_int   = filing["cik_int"]
        acc_clean = filing["accession_clean"]
        acc_d     = filing["accession_dashed"]
        primary   = filing.get("primary_doc", "")
        hdrs      = {"User-Agent": "investing-collector/1.0 contact@example.com"}
        base      = f"https://www.sec.gov/Archives/edgar/data/{cik_int}/{acc_clean}"

        # Candidates to try in order
        candidates = []
        # 1. Standard name used by most filers
        candidates.append("informationtable.xml")
        # 2. If primaryDocument is something like "0000950123-24-012345.xml" that's the cover;
        #    the info table is often adjacent with a predictable name
        if primary and primary.lower().endswith(".xml") and "primary" not in primary.lower():
            candidates.append(primary)
        # 3. Other common names
        for name in ["InfoTable.xml", "infotable.xml", "form13fInfoTable.xml",
                     "XML_Infotable.xml", "50240.xml"]:
            if name not in candidates:
                candidates.append(name)

        xml_text = None
        used_name = None
        for name in candidates:
            try:
                url = f"{base}/{name}"
                rx = self.session.get(url, timeout=20, headers=hdrs)
                if rx.status_code == 200 and len(rx.text) > 500:
                    xml_text  = rx.text
                    used_name = name
                    log.info(f"    XML found: {name} ({len(rx.text):,} chars)")
                    break
            except Exception:
                continue

        # 4. Last resort: parse index HTML for any XML link
        if not xml_text:
            try:
                idx_url = f"https://www.sec.gov/Archives/edgar/data/{cik_int}/{acc_clean}/{acc_d}-index.htm"
                rj = self.session.get(idx_url, timeout=15, headers=hdrs)
                rj.raise_for_status()
                import re as _re
                links = _re.findall(r'href="(/Archives/edgar/data/[^"]+\.xml)"', rj.text, _re.I)
                for link in links:
                    name = link.split("/")[-1]
                    if "primary" not in name.lower() and "xslForm" not in link:
                        try:
                            rx = self.session.get(f"https://www.sec.gov{link}", timeout=20, headers=hdrs)
                            if rx.status_code == 200 and len(rx.text) > 500:
                                xml_text  = rx.text
                                used_name = name
                                log.info(f"    XML found via index: {name}")
                                break
                        except Exception:
                            continue
            except Exception as e:
                log.warning(f"  EDGAR index parse: {e}")

        if not xml_text:
            log.warning(f"  EDGAR: could not retrieve XML for {acc_d}")
            return []

        results = self._parse(xml_text)
        log.info(f"    Matched: {len(results)} watched positions")
        return results


    def _parse(self, xml: str) -> list:
        """
        Parse 13F information table XML.
        Strips all namespace prefixes then uses ElementTree findall(".//infoTable").
        Matches holdings by company name via NAME_HINTS (no <ticker> tag in official spec).
        """
        import xml.etree.ElementTree as _ET
        import re as _re

        NAME_HINTS = {
            # Multiple name variants to match different filers' naming conventions
            "NVDA":  ["NVIDIA"],
            "AAPL":  ["APPLE INC", "APPLE"],
            "MSFT":  ["MICROSOFT"],
            "GOOGL": ["ALPHABET", "GOOGLE"],
            "META":  ["META PLATFORMS", "META PLATFORM", "FACEBOOK"],
            "AMZN":  ["AMAZON"],
            "TSLA":  ["TESLA"],
            "AVGO":  ["BROADCOM"],
            "TSM":   ["TAIWAN SEMI", "TAIWAN SEMICONDUCTOR"],
            "AMD":   ["ADVANCED MICRO DEVICES", "ADVANCED MICRO"],
            "WMT":   ["WALMART"],
            "PG":    ["PROCTER", "PROCTER & GAMBLE", "PROCTER AND GAMBLE"],
            "UNH":   ["UNITEDHEALTH", "UNITED HEALTH"],
            "LLY":   ["ELI LILLY", "LILLY"],
            "JPM":   ["JPMORGAN", "JP MORGAN", "JPMORGAN CHASE"],
            "BAC":   ["BANK OF AMERICA", "BANK OF AMER"],
            "CAT":   ["CATERPILLAR"],
            "HON":   ["HONEYWELL"],
            "XOM":   ["EXXON", "EXXONMOBIL", "EXXON MOBIL"],
            "CVX":   ["CHEVRON"],
            "LIN":   ["LINDE"],
            "SHW":   ["SHERWIN", "SHERWIN-WILLIAMS", "SHERWIN WILLIAMS"],
            "PLD":   ["PROLOGIS"],
            "AMT":   ["AMERICAN TOWER"],
            "NEE":   ["NEXTERA", "NEXTERA ENERGY"],
            "DUK":   ["DUKE ENERGY", "DUKE ENGY"],
        }

        # Handle namespaces: try proper ns approach first (from SEC docs),
        # fall back to stripping for robustness
        import xml.etree.ElementTree as _ET2
        # Try proper namespace method first
        try:
            root2  = _ET2.fromstring(xml)
            ns_uri = root2.tag.split("}")[0].strip("{") if "}" in root2.tag else ""
            tag    = lambda t: f"{{{ns_uri}}}{t}" if ns_uri else t
            rows2  = []
            for info in root2.iter(tag("infoTable")):
                name_  = (info.findtext(tag("nameOfIssuer")) or "").upper().strip()
                value_ = (info.findtext(tag("value")) or "0").replace(",","")
                sshr_  = info.find(f".//{tag('sshPrnamt')}")
                shares_= (sshr_.text if sshr_ is not None else "0").replace(",","")
                for ticker, hints in NAME_HINTS.items():
                    if any(h.upper() in name_ for h in hints):
                        try:
                            rows2.append({"ticker":ticker,
                                          "company":(info.findtext(tag("nameOfIssuer")) or "").strip(),
                                          "shares_k":round(float(shares_)/1000,0),
                                          "value_mn":round(float(value_)/1000,1)})
                        except Exception: pass
                        break
            if rows2:
                log.info(f"    _parse (ns method): {len(rows2)} holdings")
                return rows2
        except Exception:
            pass  # fall through to strip method

        # Strip ALL namespace declarations: xmlns="...", xmlns:ns2="...", etc.
        xml_clean = _re.sub(r'\s+xmlns(?::[a-zA-Z0-9_]+)?="[^"]*"', '', xml)
        # Strip all namespace prefixes from tags: <ns2:foo> → <foo>, </ns2:foo> → </foo>
        xml_clean = _re.sub(r'<(/?)[a-zA-Z0-9_]+:', r'<\1', xml_clean)

        rows = []
        try:
            root = _ET.fromstring(xml_clean)
            for info in root.findall(".//infoTable"):
                name   = (info.findtext("nameOfIssuer") or "").upper().strip()
                value  = (info.findtext("value") or "0").replace(",", "")
                sshr   = info.find(".//sshPrnamt")
                shares = (sshr.text if sshr is not None else "0").replace(",", "")
                for ticker, hints in NAME_HINTS.items():
                    if any(h.upper() in name for h in hints):
                        try:
                            rows.append({
                                "ticker":   ticker,
                                "company":  (info.findtext("nameOfIssuer") or "").strip(),
                                "shares_k": round(float(shares) / 1000, 0),
                                "value_mn": round(float(value)  / 1000, 1),
                            })
                        except Exception:
                            pass
                        break
        except Exception as e:
            log.warning(f"  13F XML parse error: {e}")
        log.info(f"    _parse found {len(rows)} matching holdings")
        return rows

    def collect(self) -> list:
        log.info("Fetching SEC 13F filings from EDGAR...")
        all_rows = []
        for fund, cik in self.FUNDS.items():
            log.info(f"  {fund} (CIK {cik})...")
            filing = self._latest_13f(cik)
            if not filing:
                time.sleep(1); continue
            log.info(f"    Filing: {filing['accession_dashed']}  filed: {filing['filed']}")
            holdings = self._holdings(filing)
            log.info(f"    Matched: {len(holdings)} watched positions")
            for h in holdings:
                all_rows.append({"fund": fund, "filed": filing["filed"], **h})
            time.sleep(0.5)
        log.info(f"  13F total: {len(all_rows)} positions")
        return all_rows



class MomentumScorer:
    def score(self, fred, signals, state, vix_val=None):
        pts, detail, prev_d, trend_d = {}, {}, {}, {}

        def _last(k):
            s = fred.get(k)
            return float(s.iloc[-1]) if s is not None and not s.empty else None
        def _prev(k):
            s = fred.get(k)
            return float(s.iloc[-2]) if s is not None and len(s)>=2 else None
        def _tr(n,p):
            if n is None or p is None: return "n/a"
            return "up" if n>p else ("dn" if n<p else "=")

        v=_last("t10y2y_spread"); p=_prev("t10y2y_spread")
        if v is not None:
            pts["Yield curve"]=20 if v>0.5 else 12 if v>0 else 5 if v>-0.5 else 0
            detail["Yield curve"]=f"{v:+.2f}%"
            prev_d["Yield curve"]=f"{p:+.2f}%" if p else ""
            trend_d["Yield curve"]=_tr(v,p)

        # Real VIX from yfinance ^VIX (passed in as vix_val)
        vix_now  = vix_val
        vix_prev = state.get("vix_prev")
        if vix_now is not None:
            pts["VIX"]=20 if vix_now<15 else 15 if vix_now<20 else 10 if vix_now<25 else 5 if vix_now<30 else 0
            detail["VIX"]=f"{vix_now:.2f}"
            prev_d["VIX"]=f"{vix_prev:.2f}" if vix_prev else "n/a (first run)"
            trend_d["VIX"]=_tr(vix_now,vix_prev)

        v=_last("core_pce_yoy"); p=_prev("core_pce_yoy")
        if v is not None:
            diff=abs(v*100-2.0)
            pts["Core PCE"]=20 if diff<0.3 else 15 if diff<0.5 else 10 if diff<1.0 else 5 if diff<1.5 else 0
            detail["Core PCE"]=f"{v*100:.2f}%"
            prev_d["Core PCE"]=f"{p*100:.2f}%" if p else ""
            trend_d["Core PCE"]=_tr(v,p)

        v=_last("unemployment"); p=_prev("unemployment")
        if v is not None:
            pts["Unemployment"]=20 if v<4.0 else 15 if v<5.0 else 8 if v<6.0 else 0
            detail["Unemployment"]=f"{v:.1f}%"
            prev_d["Unemployment"]=f"{p:.1f}%" if p else ""
            trend_d["Unemployment"]=_tr(v,p)

        v=_last("credit_spread_hy"); p=_prev("credit_spread_hy")
        if v is not None:
            pts["HY spread"]=20 if v<3.0 else 15 if v<4.0 else 8 if v<5.0 else 3 if v<7.0 else 0
            detail["HY spread"]=f"{v:.2f}%"
            prev_d["HY spread"]=f"{p:.2f}%" if p else ""
            trend_d["HY spread"]=_tr(v,p)

        total=sum(pts.values()); max_=len(pts)*20
        score=round(total/max_*100) if max_ else None
        regime=("Risk-ON  ok" if score and score>=70 else "Neutral  warn" if score and score>=40 else "Risk-OFF  x") if score else "N/A"

        new_state=dict(state)
        if vix_now: new_state["vix_prev"]=vix_now

        return {"score":score,"regime":regime,"pts":pts,"detail":detail,
                "prev":prev_d,"trend":trend_d,
                "as_of":datetime.today().strftime("%Y-%m-%d"),
                "new_state":new_state}


# EXCEL WRITER
class WorkbookWriter:
    SHEETS=[
        "Momentum","US Macro",
        "Indices","Signals","Forex Crypto",
        "Sectors","US Stocks","TW Stocks",
        "TW Flows","US 13F","Config",
    ]

    def __init__(self,path):
        self.path=path

    def _wb(self):
        if self.path.exists():
            wb=load_workbook(self.path, keep_links=False)
            for name in list(wb.sheetnames):
                if name not in self.SHEETS:
                    del wb[name]
            return wb
        wb=Workbook(); wb.remove(wb.active); return wb

    def _ws(self, wb, name):
        """
        Return worksheet, clearing all rows EXCEPT historical date columns.
        Date columns (header = YYYY-MM-DD) are saved and restored after clearing.
        This preserves the daily history that _append_date_columns builds up.
        """
        if name not in wb.sheetnames:
            return wb.create_sheet(name)

        ws = wb[name]

        # ── Sheets that DON'T use date columns — clear normally ──
        NO_DATE_COLS = {"US Macro", "TW Flows", "US 13F", "Momentum", "Config"}
        if name in NO_DATE_COLS:
            ws.delete_rows(1, ws.max_row)
            return ws

        # ── Forex Crypto: validate layout before trusting saved history ──
        # If the sheet was saved with the old broken layout (dates on row 1),
        # clear everything and start fresh — history from bad layout is invalid.
        if name == "Forex Crypto":
            row3_headers = [ws.cell(3, c).value for c in range(1, 7)]
            expected = ["Pair", "Rate", "1W %", "1M %", "YTD %", "As Of"]
            if row3_headers != expected:
                log.info(f"  Forex Crypto: old layout detected — forcing full clear")
                ws.delete_rows(1, ws.max_row)
                return ws

        # ── Sheets WITH date columns — save history before clearing ──
        # Some sheets have date headers on row 1 (title merged), some on row 2 or 3.
        # Forex Crypto uses row 3. Detect properly: scan rows 1-5 for YYYY-MM-DD values,
        # but skip row 1 if it's a title (the whole row is one merged dark-blue cell).
        date_header_row = None
        for hr in range(1, min(6, ws.max_row + 1)):
            found = False
            for c in range(1, ws.max_column + 1):
                v = ws.cell(hr, c).value
                if v and str(v)[:4].isdigit() and len(str(v)) == 10 and "-" in str(v):
                    # Make sure this looks like a date (YYYY-MM-DD) not a number
                    try:
                        from datetime import datetime as _dt
                        _dt.strptime(str(v)[:10], "%Y-%m-%d")
                        date_header_row = hr
                        found = True
                        break
                    except ValueError:
                        pass
            if found:
                break

        if not date_header_row:
            # No date columns yet — clear normally
            ws.delete_rows(1, ws.max_row)
            return ws

        # Save date columns keyed by TICKER (col-1 value), not row number.
        # Ticker-keyed history survives sheet rewrites where sector headers shift rows.
        last_fixed_col = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(date_header_row, c).value
            if v and not (str(v)[:4].isdigit() and len(str(v)) == 10 and "-" in str(v)):
                last_fixed_col = c
        # Build row→ticker map
        ticker_at_row = {}
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v and str(v).strip():
                ticker_at_row[r] = str(v).strip()
        # Save as {date: {ticker: value}}
        saved_dates = {}
        for c in range(last_fixed_col + 1, ws.max_column + 1):
            hval = ws.cell(date_header_row, c).value
            if not hval: continue
            try:
                from datetime import datetime as _dt2
                _dt2.strptime(str(hval)[:10], "%Y-%m-%d")
            except (ValueError, TypeError):
                continue
            date_str = str(hval)[:10]
            col_data = {ticker: ws.cell(r, c).value
                        for r, ticker in ticker_at_row.items()
                        if ws.cell(r, c).value is not None}
            if col_data:
                saved_dates[date_str] = col_data

        # Clear the sheet
        ws.delete_rows(1, ws.max_row)

        # Store saved history on the workbook for restoration after sheet rewrite
        if not hasattr(wb, '_saved_history'):
            wb._saved_history = {}
        wb._saved_history[name] = {
            "header_row": date_header_row,
            "dates": saved_dates,
        }
        log.info(f"  _ws: saved {len(saved_dates)} date cols from {name} for restoration")

        return ws

    def _momentum(self,wb,sc,fred,signals):
        """Combined dashboard + momentum scorecard. F column is a blank visual gap."""
        ws=self._ws(wb,"Momentum")

        # Column layout:
        # A-E  = Key Rates & Signals (left block)
        # F    = blank spacer
        # G-K  = Momentum Scorecard (right block)

        # ── Title: A-E only (left block width), right block has its own header ──
        ws.merge_cells("A1:E1")
        ws["A1"]=f"INVESTING DASHBOARD  |  Updated: {datetime.today().strftime('%Y-%m-%d %H:%M')}  |  FRED + Polygon.io + TWSE"
        ws["A1"].font=Font(bold=True,size=12,color="FFFFFF")
        ws["A1"].fill=_DARK; ws["A1"].alignment=_CTR

        # ── Score callout rows 3-4 (left block A-C) ──
        ws["A3"]="Momentum Score"; _hdr(ws["A3"])
        ws["B3"]="Regime"; _hdr(ws["B3"])
        ws["C3"]="As Of"; _hdr(ws["C3"])
        ws["A4"]=sc.get("score","N/A")
        ws["A4"].font=Font(bold=True,size=28,color="1F3864"); ws["A4"].alignment=_CTR
        ws.row_dimensions[4].height=40
        ws["B4"]=sc.get("regime","N/A")
        ws["B4"].font=Font(bold=True,size=13); ws["B4"].alignment=_CTR
        ws["C4"]=sc.get("as_of",""); ws["C4"].alignment=_CTR
        # PAD anchors — outside both blocks
        ws["M1"]=sc.get("score",""); ws["M1"].font=_GRN
        ws["M2"]=sc.get("regime",""); ws["M2"].font=_GRN

        # ── Left block: Key Rates (A-E, rows 6 onwards) ──
        ws.merge_cells("A6:E6")
        ws["A6"]="KEY RATES & SIGNALS"; _hdr(ws["A6"])
        for c,h in enumerate(["Indicator","Latest","Prev","Trend","Source"],1):
            _sub(ws.cell(row=7,column=c,value=h))
        inds=[
            ("Fed Funds Rate",  "fed_funds_rate",  "FEDFUNDS"),
            ("10Y Treasury",    "us_10y_yield",    "DGS10"),
            ("2Y Treasury",     "us_2y_yield",     "DGS2"),
            ("10Y-2Y Spread",   "t10y2y_spread",   "T10Y2Y"),
            ("Core PCE YoY",    "core_pce_yoy",    "PCEPILFE"),
            ("CPI YoY",         "cpi_yoy",         "CPIAUCSL"),
            ("Unemployment",    "unemployment",    "UNRATE"),
            ("HY Spread",       "credit_spread_hy","BAMLH0A0HYM2"),
        ]
        for i,(label,key,source) in enumerate(inds):
            r=8+i; alt=i%2==0
            s=fred.get(key)
            v =float(s.iloc[-1]) if s is not None and not s.empty else None
            v2=float(s.iloc[-2]) if s is not None and len(s)>=2 else None
            tr=("up" if v>v2 else "dn" if v<v2 else "=") if v and v2 else ""
            for c,val in enumerate([label,v,v2,tr,source],1):
                cell=ws.cell(row=r,column=c,value=val); _dat(cell,alt)
                if c in (2,3) and val is not None: cell.number_format="0.00"
        # VIX row — real ^VIX from yfinance, passed via sc dict
        vix_val = sc.get("vix_val")
        r = 8 + len(inds)
        for c, val in enumerate(["VIX (CBOE ^VIX)", vix_val, None, "", "yfinance ^VIX"], 1):
            _dat(ws.cell(row=r, column=c, value=val), alt=True)
        if vix_val is not None:
            ws.cell(row=r, column=2).number_format = "0.00"

        # ── Right block: Scorecard (G-K, rows 6 onwards) ──
        # G=7, H=8, I=9, J=10, K=11
        ws.merge_cells("G1:K1")
        ws["G1"]="MOMENTUM SCORECARD"; _hdr(ws["G1"])
        ws.merge_cells("G6:K6")
        ws["G6"]="SCORECARD"; _hdr(ws["G6"])
        for c,h in enumerate(["Signal","Score(/20)","Latest","Prev","Trend"],1):
            _sub(ws.cell(row=7,column=6+c,value=h))  # 6+1=7=G, 6+2=8=H, ...
        rules={
            "Yield curve": ">0.5=20|0~0.5=12|-0.5~0=5|<-0.5=0",
            "VIX":         "<15=20|15~20=15|20~25=10|25~30=5|>30=0",
            "Core PCE":    "diff from 2%: <0.3=20|<0.5=15|<1=10|<1.5=5",
            "Unemployment":"<4%=20|4~5%=15|5~6%=8|>6%=0",
            "HY spread":   "<3%=20|3~4%=15|4~5%=8|5~7%=3|>7%=0",
        }
        for i,(sig,pts) in enumerate(sc.get("pts",{}).items()):
            r=8+i; alt=i%2==0
            for c,v in enumerate([sig,pts,
                                   sc["detail"].get(sig,""),
                                   sc.get("prev",{}).get(sig,""),
                                   sc.get("trend",{}).get(sig,"")],1):
                cell=ws.cell(row=r,column=6+c,value=v); _dat(cell,alt)

        # ── Scoring rules (below scorecard) ──
        rr=8+len(sc.get("pts",{}))+1
        ws.merge_cells(f"G{rr}:K{rr}")
        ws[f"G{rr}"]="SCORING RULES"; _sec(ws[f"G{rr}"])
        for i,(sig,rule) in enumerate(rules.items()):
            r2=rr+1+i
            ws.cell(row=r2,column=7,value=sig).font=Font(bold=True)
            ws.cell(row=r2,column=8,value=rule)
            ws.merge_cells(start_row=r2,start_column=8,end_row=r2,end_column=11)

        # ── Column widths: F=blank spacer ──
        for col,w in zip("ABCDE",[26,10,10,7,18]):
            ws.column_dimensions[col].width=w
        ws.column_dimensions["F"].width=4   # blank gap — no data ever written here
        for col,w in zip("GHIJK",[18,10,14,10,8]):
            ws.column_dimensions[col].width=w

    def _macro_sheet(self,wb,sheet_name,title,fred,sections):
        ws=self._ws(wb,sheet_name)
        ws.merge_cells("A1:Z1")
        ws["A1"]=title; _hdr(ws["A1"])
        # Build monthly date index — only months where at least one series has data
        all_periods=set()
        for _,rows in sections:
            for _,k,_ in rows:
                if k.startswith("__"): continue
                s=fred.get(k)
                if s is not None and not s.empty:
                    s2=s.copy(); s2.index=pd.to_datetime(s2.index)
                    # Resample to monthly — only add months with actual values
                    monthly=s2.resample("MS").last().dropna()
                    for d in monthly.index:
                        all_periods.add((d.year,d.month))
        date_index=sorted(all_periods)[-HISTORY_MONTHS:]
        DC=3
        _sub(ws.cell(row=2,column=1,value="Section"))
        _sub(ws.cell(row=2,column=2,value="Indicator"))
        for i,(yr,mo) in enumerate(date_index):
            _sub(ws.cell(row=2,column=DC+i,value=f"{yr}-{mo:02d}"))
        row=3
        for sec_name,indicators in sections:
            c1=ws.cell(row=row,column=1,value=sec_name); _sec(c1)
            end_col=max(DC+len(date_index)-1,DC)
            ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=end_col)
            row+=1
            for label,key,_ in indicators:
                alt=row%2==0; is_manual=key.startswith("__")
                _dat(ws.cell(row=row,column=1,value=""),alt)
                _dat(ws.cell(row=row,column=2,value=label),alt,yellow=is_manual)
                if not is_manual:
                    s=fred.get(key)
                    if s is not None and not s.empty:
                        s.index=pd.to_datetime(s.index)
                        monthly=s.resample("MS").last()
                        for i,(yr,mo) in enumerate(date_index):
                            val=next((round(float(v),4) for ts,v in monthly.items()
                                      if ts.year==yr and ts.month==mo),None)
                            cell=ws.cell(row=row,column=DC+i,value=val)
                            _dat(cell,alt)
                            if val is not None: cell.number_format="0.00"
                else:
                    for i in range(len(date_index)):
                        _dat(ws.cell(row=row,column=DC+i,value=None),alt,yellow=True)
                row+=1
        ws.freeze_panes="C3"
        ws.column_dimensions["A"].width=12
        ws.column_dimensions["B"].width=28

    def _indices_sheet(self, wb, df: pd.DataFrame):
        ws = self._ws(wb, "Indices")
        n_cols = 7
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        ws["A1"] = "GLOBAL MARKET INDICES  —  yfinance / TWSE / Stooq / Polygon ETF proxies"
        _hdr(ws["A1"])

        headers = ["Ticker", "Index", "Last Value", "1W %", "1M %", "YTD %", "As Of"]
        for c, h in enumerate(headers, 1):
            _sub(ws.cell(row=2, column=c, value=h))

        if df is None or df.empty:
            ws.cell(row=3, column=1, value="No index data — check yfinance / network.")
            return

        # Region labels
        REGIONS = {
            "^GSPC": "── US ──", "^TWII": "── Taiwan ──",
            "^N225": "── Asia ──", "^FTSE": "── Europe ──",
        }

        ri = 3
        for _, row in df.iterrows():
            # Region label row
            if row["ticker"] in REGIONS:
                ws.merge_cells(f"A{ri}:{get_column_letter(n_cols)}{ri}")
                label_cell = ws[f"A{ri}"]
                label_cell.value = REGIONS[row["ticker"]]
                label_cell.font = Font(bold=True, color="1F3864", size=10)
                label_cell.fill = PatternFill("solid", fgColor="D9E1F2")
                label_cell.alignment = Alignment(horizontal="center")
                ri += 1

            alt = (ri % 2 == 0)
            vals = [
                row.get("ticker"), row.get("name"),
                row.get("last_price"), row.get("chg_1w_pct"),
                row.get("chg_1m_pct"), row.get("chg_ytd_pct"),
                str(row.get("as_of", ""))[:10],
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                _dat(cell, alt)
                if ci == 3 and val is not None:
                    cell.number_format = "#,##0.00"
                if ci in (4, 5, 6) and val is not None:
                    cell.number_format = "0.00"
                    try:
                        v = float(val)
                        cell.font = Font(
                            color="008000" if v > 0 else ("C00000" if v < 0 else "000000")
                        )
                    except: pass
            ri += 1
        _autofit(ws)


    def _price_sheet(self,wb,sheet_name,title,df):
        ws=self._ws(wb,sheet_name)
        headers=["Ticker","Name","Last Price","1W %","1M %","YTD %",
                 "52W Hi","52W Lo","% from Hi","As Of"]
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"]=title; _hdr(ws["A1"])
        for c,h in enumerate(headers,1):
            _sub(ws.cell(row=2,column=c,value=h))
        if df.empty:
            ws.cell(row=3,column=1,value="No data.")
        else:
            for i,r in df.reset_index(drop=True).iterrows():
                alt=i%2==0
                for c,v in enumerate([r.get("ticker"),r.get("name"),r.get("last_price"),
                                       r.get("chg_1w_pct"),r.get("chg_1m_pct"),r.get("chg_ytd_pct"),
                                       r.get("high_52w"),r.get("low_52w"),r.get("pct_from_hi"),r.get("as_of")],1):
                    cell=ws.cell(row=i+3,column=c,value=v); _dat(cell,alt)
                    if c in (4,5,6,9) and v is not None: cell.number_format="0.00"
                    if c==3 and v is not None: cell.number_format="#,##0.00"
                    if c in (10,11) and v is not None: cell.number_format="0.00"
                    if c==12 and v is not None: cell.number_format="0.00"
        _autofit(ws)

    # ── Forex Crypto sheet column layout (fixed, never changes) ──────────────
    # Row 1  : title bar  A1:J1
    # Row 2  : "FOREX RATES" section bar  A2:F2
    # Row 3  : forex column headers  A-F  [Pair | Rate | 1W% | 1M% | YTD% | As Of]
    # Rows 4+ : forex data
    # Gap row : blank
    # "CRYPTOCURRENCY" section bar  A:J
    # Crypto column headers  A-J  [Ticker|Name|Last|1W%|1M%|YTD%|52Hi|52Lo|%Hi|AsOf]
    # Crypto data rows
    # Columns G+ : date-column history appended by _append_date_columns
    #
    # IMPORTANT: _append_date_columns uses row_for_header=3 for this sheet
    # so history date headers land on row 3, aligned with forex headers.

    _FC_FOREX_COLS   = 6   # A-F  (Pair, Rate, 1W%, 1M%, YTD%, As Of)
    _FC_CRYPTO_COLS  = 10  # A-J

    def _forex_crypto_sheet(self, wb, forex_df, crypto_df):
        ws = self._ws(wb, "Forex Crypto")

        # ── Row 1: Title (span max of forex/crypto cols) ──
        title_end = get_column_letter(self._FC_CRYPTO_COLS)
        ws.merge_cells(f"A1:{title_end}1")
        ws["A1"] = "FOREX RATES  --  yfinance (primary)  /  Polygon.io (fallback)"
        _hdr(ws["A1"])

        # ── Row 2: Forex section bar ──
        ws.merge_cells(f"A2:{get_column_letter(self._FC_FOREX_COLS)}2")
        _sec(ws["A2"]); ws["A2"].value = "FOREX RATES"

        # ── Row 3: Forex column headers  (this is the history anchor row) ──
        for c, h in enumerate(["Pair", "Rate", "1W %", "1M %", "YTD %", "As Of"], 1):
            _sub(ws.cell(row=3, column=c, value=h))

        # ── Rows 4+: Forex data ──
        forex_end = 3
        if forex_df is not None and not forex_df.empty:
            for i, r in forex_df.reset_index(drop=True).iterrows():
                row = i + 4; alt = i % 2 == 0; forex_end = row
                vals = [r.get("pair"), r.get("last_price"),
                        r.get("chg_1w_pct"), r.get("chg_1m_pct"),
                        r.get("chg_ytd_pct"), str(r.get("as_of") or "")[:10]]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=c, value=v); _dat(cell, alt)
                    if c == 2 and v is not None: cell.number_format = "0.0000"
                    if c in (3, 4, 5) and v is not None: cell.number_format = "0.00"

        # ── Gap + Crypto section ──
        offset = forex_end + 2
        ws.merge_cells(f"A{offset}:{get_column_letter(self._FC_CRYPTO_COLS)}{offset}")
        _sec(ws.cell(row=offset, column=1))
        ws.cell(row=offset, column=1).value = "CRYPTOCURRENCY"

        for c, h in enumerate(["Ticker", "Name", "Last Price", "1W %", "1M %",
                                "YTD %", "52W Hi", "52W Lo", "% from Hi", "As Of"], 1):
            _sub(ws.cell(row=offset + 1, column=c, value=h))

        if crypto_df is not None and not crypto_df.empty:
            for i, r in crypto_df.reset_index(drop=True).iterrows():
                row = offset + 2 + i; alt = i % 2 == 0
                for c, v in enumerate([
                    r.get("ticker"), r.get("name"), r.get("last_price"),
                    r.get("chg_1w_pct"), r.get("chg_1m_pct"), r.get("chg_ytd_pct"),
                    r.get("high_52w"), r.get("low_52w"), r.get("pct_from_hi"),
                    str(r.get("as_of") or "")[:10]
                ], 1):
                    cell = ws.cell(row=row, column=c, value=v); _dat(cell, alt)
                    if c == 3 and v is not None: cell.number_format = "#,##0.00"
                    if c in (4, 5, 6, 9) and v is not None: cell.number_format = "0.00"
                    if c in (7, 8) and v is not None: cell.number_format = "#,##0.00"

        _autofit(ws)

    def _stocks_by_sector(self,wb,sheet_name,title,df):
        ws=self._ws(wb,sheet_name)
        headers=["Ticker","Name","Last Price","1W %","1M %","YTD %",
                 "52W Hi","52W Lo","% from Hi","P/E","P/B","Div Yield %","As Of"]
        n=len(headers)
        ws.merge_cells(f"A1:{get_column_letter(n)}1")
        ws["A1"]=title; _hdr(ws["A1"])
        if df.empty:
            ws.cell(row=2,column=1,value="No data."); _autofit(ws); return
        row=2
        for sector in df["sector"].unique():
            c1=ws.cell(row=row,column=1,value=sector); _sec(c1)
            ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
            row+=1
            for c,h in enumerate(headers,1):
                _sub(ws.cell(row=row,column=c,value=h))
            row+=1
            for i,r in df[df["sector"]==sector].reset_index(drop=True).iterrows():
                alt=i%2==0
                for c,v in enumerate([r.get("ticker"),r.get("name"),r.get("last_price"),
                                       r.get("chg_1w_pct"),r.get("chg_1m_pct"),r.get("chg_ytd_pct"),
                                       r.get("high_52w"),r.get("low_52w"),r.get("pct_from_hi"),
                                       r.get("pe_ratio"),r.get("pb_ratio"),r.get("div_yield"),r.get("as_of")],1):
                    cell=ws.cell(row=row,column=c,value=v); _dat(cell,alt)
                    if c in (4,5,6,9) and v is not None: cell.number_format="0.00"
                    if c==3 and v is not None: cell.number_format="#,##0.00"
                    if c in (10,11) and v is not None: cell.number_format="0.00"
                    if c==12 and v is not None: cell.number_format="0.00"
                    if c==12 and v is not None: cell.number_format="0.00"
                row+=1
            row+=1
        _autofit(ws)

    def _tw_stocks(self, wb, df: pd.DataFrame):
        ws = self._ws(wb, "TW Stocks")
        headers = ["Ticker","Name","Last Price","Change","Chg %",
                   "Open","High","Low","Volume (K)","P/E","P/B","Div Yield %","As Of"]
        n = len(headers)

        ws.merge_cells(f"A1:{get_column_letter(n)}1")
        ws["A1"] = "TAIWAN STOCKS  --  TWSE OpenAPI (free, no key)"
        _hdr(ws["A1"])

        if df.empty or df["last_price"].isna().all():
            ws.cell(row=2, column=1,
                    value="No data — TWSE market closed or API temporarily unavailable.")
            _autofit(ws); return

        row = 2
        for sector in df["sector"].unique():
            sdf = df[df["sector"]==sector].reset_index(drop=True)
            # Sector header row
            c1 = ws.cell(row=row, column=1, value=sector); _sec(c1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
            row += 1
            # Column headers
            for c, h in enumerate(headers, 1):
                _sub(ws.cell(row=row, column=c, value=h))
            row += 1
            # Data rows
            for i, r in sdf.iterrows():
                alt = i % 2 == 0
                chg_pct = r.get("chg_pct")
                for c, v in enumerate([
                    r.get("ticker"), r.get("name"), r.get("last_price"),
                    r.get("change"), chg_pct,
                    r.get("open"), r.get("high"), r.get("low"),
                    r.get("volume_k"), r.get("pe_ratio"), r.get("pb_ratio"),
                    r.get("div_yield"), r.get("as_of")
                ], 1):
                    cell = ws.cell(row=row, column=c, value=v); _dat(cell, alt)
                    if c in (3,6,7,8) and v is not None: cell.number_format = "#,##0.00"
                    if c in (4,5,11,12) and v is not None: cell.number_format = "0.00"
                    if c == 10 and v is not None: cell.number_format = "0.00"
                    # Color code the change column
                    if c == 4 and v is not None:
                        cell.font = Font(color="CC0000", bold=True) if v < 0 else Font(color="008000", bold=True) if v > 0 else _BLK
                    if c == 5 and v is not None:
                        cell.font = Font(color="CC0000", bold=True) if v < 0 else Font(color="008000", bold=True) if v > 0 else _BLK
                row += 1
            row += 1  # spacer

        _autofit(ws)

    def _tw_flows(self,wb,flows,stock_flows=None):
        ws=self._ws(wb,"TW Flows")

        # ── Section 1: Market total ──
        ws.merge_cells("A1:E1")
        ws["A1"]=f"TAIWAN 三大法人  --  TWSE  |  As of: {flows.get('as_of','N/A')}"; _hdr(ws["A1"])
        ws.merge_cells("A2:E2")
        c=ws["A2"]; c.value="MARKET TOTAL (大盤)"; _sec(c)
        for c,h in enumerate(["Entity","Buy (億 TWD)","Sell (億 TWD)","Net (億 TWD)","Direction"],1):
            _sub(ws.cell(row=3,column=c,value=h))
        mkt_rows=flows.get("rows",[])
        data_end=3
        if not mkt_rows:
            ws.cell(row=4,column=1,value="No data -- TWSE unavailable (weekend/holiday).")
            data_end=5
        else:
            for i,r in enumerate(mkt_rows):
                row=4+i; alt=i%2==0; data_end=row
                net=r.get("net_bn",0) or 0
                direction="IN" if net>0 else ("OUT" if net<0 else "=")
                for c,v in enumerate([r.get("entity"),r.get("buy_bn"),r.get("sell_bn"),net,direction],1):
                    cell=ws.cell(row=row,column=c,value=v); _dat(cell,alt)
                    if c in (2,3,4) and v is not None: cell.number_format="#,##0.00"
                    if c==5:
                        if direction=="IN":  cell.font=Font(color="008000",bold=True)
                        elif direction=="OUT": cell.font=_RED

        # ── Section 2: Per-stock breakdown ──
        # Only write if we have actual flow data (T86 returned rows)
        has_flow_data = stock_flows and any(
            sf.get("total_net") is not None for sf in stock_flows
        )
        if has_flow_data:
            off=data_end+2
            ws.merge_cells(f"A{off}:G{off}")
            c=ws.cell(row=off,column=1); c.value="PER-STOCK FLOWS (個股三大法人, 千股 net)"; _sec(c)
            for c,h in enumerate(["Stock","Name","Foreign Net","Trust Net","Dealer Net","Total Net","Direction"],1):
                _sub(ws.cell(row=off+1,column=c,value=h))
            name_lk={code:name for code,(sec,name) in TWStockCollector.WATCHLIST.items()}
            for i,sf in enumerate(stock_flows):
                r=off+2+i; alt=i%2==0
                sym=sf.get("stock", sf.get("ticker",""))
                sym_clean = str(sym).replace(".TW","").replace(".tw","").strip()
                tnet=sf.get("total_net") or 0
                direction="IN" if tnet>0 else ("OUT" if tnet<0 else "=")
                for c,v in enumerate([sym_clean, name_lk.get(sym_clean, sym_clean),
                                       sf.get("foreign_net"),sf.get("trust_net"),
                                       sf.get("dealer_net"),sf.get("total_net"),direction],1):
                    cell=ws.cell(row=r,column=c,value=v); _dat(cell,alt)
                    if c in (3,4,5,6) and v is not None: cell.number_format="#,##0.00"
                    if c==7:
                        if direction=="IN":  cell.font=Font(color="008000",bold=True)
                        elif direction=="OUT": cell.font=_RED
        elif stock_flows:
            log.warning("  Per-stock flows: T86 had no data — previous sheet data preserved")
        _autofit(ws)

        _autofit(ws)

    def _13f_sheet(self, wb, rows_13f: list):
        ws = self._ws(wb, "US 13F")
        ws.merge_cells("A1:G1")
        ws["A1"] = ("SEC 13F INSTITUTIONAL HOLDINGS  --  EDGAR  |  "
                    "Quarterly filing, up to 45-day lag  |  Filtered to your US watchlist")
        _hdr(ws["A1"])

        note_row = 2
        ws.merge_cells(f"A{note_row}:G{note_row}")
        ws[f"A{note_row}"] = ("Funds: Berkshire Hathaway  |  BlackRock  |  Vanguard  "
                              "|  State Street  |  Bridgewater")
        ws[f"A{note_row}"].font = Font(italic=True, size=9, color="595959")
        ws[f"A{note_row}"].alignment = _LFT

        for c, h in enumerate(["Fund","Filed Date","Ticker","Company",
                                "Shares (K)","Value (USD M)","% of Watch"], 1):
            _sub(ws.cell(row=3, column=c, value=h))

        if not rows_13f:
            ws.cell(row=4, column=1,
                    value="No 13F data — EDGAR may be temporarily unavailable, or no filings in current quarter.")
            _autofit(ws)
            return

        # Compute total value per fund for % calculation
        fund_totals = {}
        for r in rows_13f:
            fund_totals[r["fund"]] = fund_totals.get(r["fund"], 0) + (r.get("value_mn") or 0)

        # Group by fund, then by ticker
        from collections import defaultdict
        by_fund = defaultdict(list)
        for r in rows_13f:
            by_fund[r["fund"]].append(r)

        row = 4
        for fund_name, holdings in by_fund.items():
            # Fund section header
            c1 = ws.cell(row=row, column=1, value=fund_name); _sec(c1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            row += 1
            # Sort by value descending
            for i, h in enumerate(sorted(holdings, key=lambda x: x.get("value_mn") or 0, reverse=True)):
                alt = i % 2 == 0
                fund_tot = fund_totals.get(fund_name, 0)
                pct = round(h.get("value_mn", 0) / fund_tot * 100, 1) if fund_tot else None
                for c, v in enumerate([fund_name, h.get("filed"), h.get("ticker"),
                                        h.get("company"), h.get("shares_k"),
                                        h.get("value_mn"), pct], 1):
                    cell = ws.cell(row=row, column=c, value=v); _dat(cell, alt)
                    if c == 5 and v is not None: cell.number_format = "#,##0"
                    if c == 6 and v is not None: cell.number_format = "#,##0.0"
                    if c == 7 and v is not None: cell.number_format = "0.0"
                row += 1
            row += 1  # spacer

        _autofit(ws)

    # Ticker rename bridge: old ticker → new ticker for history continuity
    LEGACY_TICKER_MAP = {
        "2311":    "3711",
        "2311.TW": "3711.TW",
    }

    def _restore_history(self, wb):
        """
        Restore historical date columns saved by _ws() before sheet clearing.
        History is ticker-keyed: {date: {ticker: value}} — survives row shifts.
        LEGACY_TICKER_MAP bridges renamed tickers (e.g. 2311 → 3711).
        """
        if not hasattr(wb, '_saved_history') or not wb._saved_history:
            return

        for sheet_name, history in wb._saved_history.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws         = wb[sheet_name]
            header_row = history["header_row"]
            saved_dates = history["dates"]   # {date: {ticker: value}}
            if not saved_dates:
                continue

            # Build ticker→row from rewritten sheet col 1
            ticker_to_row = {}
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, 1).value
                if v and str(v).strip():
                    ticker_to_row[str(v).strip()] = r

            # Find first empty col after fixed columns
            last_fixed_col = 0
            for c in range(1, ws.max_column + 1):
                v = ws.cell(header_row, c).value
                if v and not (str(v)[:4].isdigit() and len(str(v)) == 10 and "-" in str(v)):
                    last_fixed_col = c
            next_col = last_fixed_col + 1

            for date_str in sorted(saved_dates.keys()):
                col_data = saved_dates[date_str]  # {ticker: value}
                if not col_data:
                    continue
                # Write date header
                coord = ws.cell(header_row, next_col).coordinate
                for mr in list(ws.merged_cells.ranges):
                    if coord in mr:
                        ws.unmerge_cells(str(mr)); break
                hc = ws.cell(header_row, next_col)
                hc.value = date_str
                hc.font      = Font(bold=True, color="FFFFFF", size=9)
                hc.fill      = PatternFill("solid", fgColor="1F3864")
                hc.alignment = Alignment(horizontal="center")
                # Write values by ticker lookup (with legacy bridge)
                for ticker, val in col_data.items():
                    resolved = self.LEGACY_TICKER_MAP.get(ticker, ticker)
                    row_num  = ticker_to_row.get(resolved)
                    if row_num is None or row_num == header_row:
                        continue
                    c = ws.cell(row_num, next_col)
                    c.value = val
                    c.font      = Font(size=9)
                    c.alignment = Alignment(horizontal="center")
                    if isinstance(val, (int, float)):
                        c.number_format = "0.00"
                next_col += 1

            log.info(f"  Restored {len(saved_dates)} date cols → {sheet_name}")

        wb._saved_history = {}

    def _append_date_columns(self, wb, sc, signals, sectors,
                              indices, forex, crypto, us_stocks,
                              tw_stocks):
        """
        Append today's values as a new date column to the RIGHT of every
        data sheet. Each run adds one column; date is the column header.
        Duplicate-date guard prevents double-writing on re-runs.
        Sheets covered: Momentum, Signals, Indices, Forex Crypto,
                        Sectors, US Stocks, TW Stocks, TW Flows
        Skipped: US Macro (already time-series), US 13F (quarterly), Config
        """
        today = datetime.today().strftime("%Y-%m-%d")
        log.info(f"  Appending historical date column for {today}...")

        def _find_or_create_col(ws, row_for_header=1):
            """
            Scan the header row left-to-right.
            Return the next empty column index.
            Return None if today's date already exists in any column (skip re-run).
            """
            nc = 1
            while True:
                cell_val = ws.cell(row=row_for_header, column=nc).value
                if cell_val is None:
                    return nc          # found first empty column
                if str(cell_val)[:10] == today:
                    log.info(f"    {ws.title}: {today} already exists at col {nc} — skipping")
                    return None        # already written today
                nc += 1

        def _write_header(ws, col, row=1):
            # Must unmerge BEFORE fetching the cell object.
            # After unmerge_cells, re-fetch to get a regular Cell (not MergedCell).
            coord = ws.cell(row=row, column=col).coordinate
            for mr in list(ws.merged_cells.ranges):
                if coord in mr:
                    ws.unmerge_cells(str(mr))
                    break
            # Re-fetch after unmerge — now guaranteed to be a writable Cell
            c = ws.cell(row=row, column=col)
            c.value     = today
            c.font      = Font(bold=True, color="FFFFFF", size=9)
            c.fill      = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(horizontal="center")

        def _write_val(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="center")
            c.font      = Font(size=9)
            if isinstance(val, float):
                c.number_format = "0.00"

        def _df_price(df, ticker_col, ticker_val, price_col="last_price"):
            """Safely get a price value from a DataFrame."""
            if df is None or (hasattr(df,'empty') and df.empty): return None
            try:
                row = df[df[ticker_col] == str(ticker_val)]
                if row.empty: return None
                v = row.iloc[0].get(price_col)
                return round(float(v), 2) if v is not None else None
            except: return None

        # ── MOMENTUM ─────────────────────────────────────────────────────────
        if "Momentum" in wb.sheetnames:
            ws  = wb["Momentum"]
            col = _find_or_create_col(ws, row_for_header=3)
            if col:
                _write_header(ws, col, row=3)
                rates = {r["indicator"]: r["latest"] for r in sc.get("rates", [])}
                _write_val(ws, 4, col, sc.get("score"))
                _write_val(ws, 5, col, str(sc.get("regime","")).replace("✅","").replace("❌","").replace("⚠️","").strip())
                for ri, key in enumerate([
                    "Fed Funds Rate","10Y Treasury","2Y Treasury","10Y-2Y Spread",
                    "CPI YoY","Core PCE YoY","Unemployment","HY Spread","VIX (CBOE ^VIX)"
                ], 7):
                    _write_val(ws, ri, col, rates.get(key))
                log.info(f"    Momentum: col {col} written")

        # ── SIGNALS ───────────────────────────────────────────────────────────
        if "Signals" in wb.sheetnames and signals is not None:
            ws  = wb["Signals"]
            col = _find_or_create_col(ws, row_for_header=2)
            if col:
                _write_header(ws, col, row=2)
                for ri in range(3, ws.max_row + 1):
                    ticker = ws.cell(row=ri, column=1).value
                    if ticker:
                        _write_val(ws, ri, col, _df_price(signals, "ticker", ticker))
                log.info(f"    Signals: col {col} written")

        # ── INDICES ───────────────────────────────────────────────────────────
        if "Indices" in wb.sheetnames and indices is not None:
            ws  = wb["Indices"]
            col = _find_or_create_col(ws, row_for_header=2)
            if col:
                _write_header(ws, col, row=2)
                for ri in range(3, ws.max_row + 1):
                    ticker = ws.cell(row=ri, column=1).value
                    if ticker and not str(ticker).startswith("─"):
                        _write_val(ws, ri, col, _df_price(indices, "ticker", ticker))
                log.info(f"    Indices: col {col} written")

        # ── FOREX CRYPTO ──────────────────────────────────────────────────────
        if "Forex Crypto" in wb.sheetnames:
            ws  = wb["Forex Crypto"]
            # Row 3 holds the forex column headers [Pair|Rate|1W%|1M%|YTD%|As Of]
            # Date columns are appended to the RIGHT of those fixed columns.
            col = _find_or_create_col(ws, row_for_header=3)
            if col:
                _write_header(ws, col, row=3)
                # Write forex prices (rows 4–7 typically; col 1 = "Pair" label like "EUR/USD")
                for ri in range(4, ws.max_row + 1):
                    pair = ws.cell(row=ri, column=1).value
                    if pair and "/" in str(pair):
                        # Look up in forex_df by pair name
                        if forex is not None and not forex.empty and "pair" in forex.columns:
                            _write_val(ws, ri, col, _df_price(forex, "pair", str(pair)))
                # Write crypto prices (look for X: tickers in col 1)
                for ri in range(4, ws.max_row + 1):
                    ticker = ws.cell(row=ri, column=1).value
                    if ticker and str(ticker).startswith("X:"):
                        if crypto is not None and not crypto.empty:
                            _write_val(ws, ri, col, _df_price(crypto, "ticker", str(ticker)))
                log.info(f"    Forex Crypto: col {col} written (header row=3)")

        # ── SECTORS ───────────────────────────────────────────────────────────
        if "Sectors" in wb.sheetnames and sectors is not None:
            ws  = wb["Sectors"]
            col = _find_or_create_col(ws, row_for_header=2)
            if col:
                _write_header(ws, col, row=2)
                for ri in range(3, ws.max_row + 1):
                    ticker = ws.cell(row=ri, column=1).value
                    if ticker:
                        _write_val(ws, ri, col, _df_price(sectors, "ticker", ticker))
                log.info(f"    Sectors: col {col} written")

        # ── US STOCKS ─────────────────────────────────────────────────────────
        if "US Stocks" in wb.sheetnames and us_stocks is not None:
            ws  = wb["US Stocks"]
            # Row 3 = "Ticker/Name/Last Price..." header; data starts row 4
            col = _find_or_create_col(ws, row_for_header=3)
            if col:
                _write_header(ws, col, row=3)
                for ri in range(4, ws.max_row + 1):
                    ticker = ws.cell(row=ri, column=1).value
                    price  = ws.cell(row=ri, column=3).value
                    # Skip blank rows and sector header rows (no numeric price)
                    if ticker and isinstance(price, (int, float)):
                        _write_val(ws, ri, col, _df_price(us_stocks, "ticker", str(ticker)))
                log.info(f"    US Stocks: col {col} written")

        # ── TW STOCKS ─────────────────────────────────────────────────────────
        if "TW Stocks" in wb.sheetnames and tw_stocks is not None:
            ws  = wb["TW Stocks"]
            # Row 3 = "Ticker/Name/Last Price..." header; data starts row 4
            col = _find_or_create_col(ws, row_for_header=3)
            if col:
                _write_header(ws, col, row=3)
                for ri in range(4, ws.max_row + 1):
                    ticker = ws.cell(row=ri, column=1).value
                    price  = ws.cell(row=ri, column=3).value
                    # Skip blank rows and sector header rows
                    if ticker and isinstance(price, (int, float)):
                        _write_val(ws, ri, col, _df_price(tw_stocks, "ticker", str(ticker)))
                log.info(f"    TW Stocks: col {col} written")


        # TW Flows intentionally excluded from date-column history.
        # It is rewritten fresh each run — appending date cols breaks the merged title.
        # Historical flow data is captured in the Market Total net values above.

        log.info(f"  Historical date columns complete for {today}")


    def _config(self,wb):
        ws=self._ws(wb,"Config")
        ws.merge_cells("A1:B1")
        ws["A1"]="CONFIGURATION"; _hdr(ws["A1"])
        for i,(k,v) in enumerate([
            ("FRED API key env var","FRED_API_KEY"),
            ("FRED key signup","https://fred.stlouisfed.org/docs/api/api_key.html"),
            ("Polygon API key env var","POLYGON_API_KEY"),
            ("Polygon key signup","https://polygon.io"),
            ("Polygon rate limit",f"{POLY_PAUSE}s between calls"),
            ("Workbook path",str(self.path)),
            ("State file (VIX prev)",str(STATE_FILE)),
            ("History months",HISTORY_MONTHS),
            ("Last run",datetime.today().strftime("%Y-%m-%d %H:%M")),
            ("PAD Score anchor","Momentum sheet  J1"),
            ("PAD Regime anchor","Momentum sheet  J2"),
            ("Yellow cells","Manual input required (PMI etc.)"),
            ("US inst. flows","No real-time equivalent to TWSE 三大法人 — 13F is quarterly (45-day lag)"),
            ("13F source","SEC EDGAR -- public, no key needed -- https://www.sec.gov/cgi-bin/browse-edgar"),
        ]):
            r=i+2
            ws.cell(row=r,column=1,value=k).font=Font(bold=True)
            ws.cell(row=r,column=2,value=v).font=_BLUE
        _autofit(ws)

    def write(self,fred_us,sc,signals,forex,crypto,sectors,us_stocks,tw_stocks,tw_flows,indices=None,stock_flows=None,rows_13f=None):
        log.info(f"Writing workbook -> {self.path}")
        wb=self._wb()
        for name in self.SHEETS:
            if name not in wb.sheetnames:
                wb.create_sheet(name)

        us_sections=[
            ("RATES",[
                ("Fed Funds Rate (%)","fed_funds_rate",False),
                ("10Y Treasury (%)","us_10y_yield",False),
                ("2Y Treasury (%)","us_2y_yield",False),
                ("3M T-Bill (%)","us_3m_tbill",False),
                ("10Y-2Y Spread (%)","t10y2y_spread",False),
                ("10Y-3M Spread (%)","t10y3m_spread",False),
            ]),
            ("INFLATION",[
                ("CPI YoY (%)","cpi_yoy",True),
                ("Core PCE YoY (%)","core_pce_yoy",True),
                ("PPI YoY (%)","ppi_yoy",True),
                ("ISM Mfg PMI  [MANUAL]","__ism_pmi__",False),
                ("ISM Svc PMI  [MANUAL]","__ism_svc__",False),
            ]),
            ("GROWTH",[
                ("Real GDP Growth (%)","gdp_growth",False),
                ("Industrial Prod YoY (%)","industrial_prod",True),
                ("Capacity Util (%)","capacity_util",False),
            ]),
            ("LABOUR",[
                ("Unemployment (%)","unemployment",False),
                ("Nonfarm Payrolls (K delta)","nonfarm_payrolls",False),
            ]),
            ("CONSUMER",[
                ("Consumer Sentiment","consumer_sentiment",False),
                ("Personal Saving (%)","personal_saving",False),
            ]),
            ("CREDIT",[
                ("HY Spread (%)","credit_spread_hy",False),
                ("IG Spread (%)","credit_spread_ig",False),
                ("M2 YoY (%)","m2_supply",True),
            ]),
        ]
        tw_sections=[
            ("GROWTH",[
                ("Real GDP (level, Penn WT)","tw_gdp",False),
                ("Exports YoY (%)","tw_exports",False),
            ]),
            ("INFLATION",[
                ("CPI Annual (%)","tw_cpi",False),
                ("Mfg PMI  [MANUAL]","__tw_pmi__",False),
            ]),
            ("LABOUR",[
                ("Unemployment (%)","tw_unemployment",False),
            ]),
        ]

        self._momentum(wb,sc,fred_us,signals)
        if indices is not None: self._indices_sheet(wb,indices)
        self._macro_sheet(wb,"US Macro","US MACRO  --  FRED",fred_us,us_sections)
        # Merge crypto into signals for unified Signals sheet
        _crypto_for_signals = crypto.copy() if (crypto is not None and not crypto.empty) else pd.DataFrame()
        if not _crypto_for_signals.empty:
            for col in ["high_52w","low_52w","pct_from_hi"]:
                if col not in _crypto_for_signals.columns:
                    _crypto_for_signals[col] = None
        signals_with_crypto = pd.concat([signals, _crypto_for_signals], ignore_index=True)             if not _crypto_for_signals.empty else signals
        self._price_sheet(wb,"Signals","MARKET SIGNALS  —  Equities · Bonds · Commodities · FX · Crypto  (Polygon / yfinance)",signals_with_crypto)
        self._forex_crypto_sheet(wb,forex,None)  # crypto moved to Signals
        self._price_sheet(wb,"Sectors","US SECTOR ETFs (SPDR)  --  Polygon.io",sectors)
        self._stocks_by_sector(wb,"US Stocks","US STOCKS -- TOP 3/SECTOR -- Polygon.io",us_stocks)
        self._tw_stocks(wb,tw_stocks)
        self._tw_flows(wb,tw_flows,stock_flows)
        self._13f_sheet(wb,rows_13f)
        # Restore saved historical date columns for all sheets
        self._restore_history(wb)

        self._append_date_columns(wb,sc,signals,sectors,
                                  indices,forex,crypto,us_stocks,
                                  tw_stocks)
        self._config(wb)

        for i,name in enumerate(self.SHEETS):
            if name in wb.sheetnames:
                idx=wb.sheetnames.index(name)
                if idx!=i: wb.move_sheet(name,offset=idx-i)

        wb.save(self.path)
        log.info(f"  Saved -> {self.path}")


# MAIN
def run():
    log.info("="*60)
    log.info("  INVESTING DATA COLLECTOR  v3.8")
    log.info(f"  {datetime.today().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("="*60)

    if "YOUR_FRED_KEY_HERE" in FRED_API_KEY:
        log.error("Set: setx FRED_API_KEY \"your_key\""); sys.exit(1)
    if "YOUR_POLYGON_KEY_HERE" in POLYGON_API_KEY:
        log.error("Set: setx POLYGON_API_KEY \"your_key\""); sys.exit(1)

    state=load_state()

    fred=FREDCollector(FRED_API_KEY)
    fred_us=fred.collect_us()
    fred_tw=fred.collect_tw_yoy()

    poly=PolygonCollector(POLYGON_API_KEY)
    signals   = poly.collect_signals()
    vix_val   = poly.collect_vix()          # real ^VIX from yfinance
    crypto    = poly.collect_crypto()
    forex     = poly.collect_forex()
    sectors   = poly.collect_sectors()
    us_stocks = poly.collect_us_stocks()
    indices   = GlobalIndexCollector().collect(poly_collector=poly)
    tw_stocks = TWStockCollector().collect()

    twse = TWInstitutionalCollector()
    tw_flows = twse.collect()
    all_tw_syms = [sym for lst in PolygonCollector.TW_STOCKS.values() for sym,_ in lst]
    stock_flows = twse.collect_stock_flows(all_tw_syms)

    vix_val = poly.collect_vix()
    sc=MomentumScorer().score(fred_us, signals, state, vix_val=vix_val)
    sc["vix_val"] = vix_val                 # carry into sheet writer
    log.info(f"  Score: {sc['score']}/100  ->  {sc['regime']}  VIX={vix_val}")

    _ns = sc.pop("new_state", dict(state))
    if vix_val is not None: _ns["vix_prev"] = vix_val
    save_state(_ns)

    rows_13f = SEC13FCollector().collect()

    WorkbookWriter(WORKBOOK_PATH).write(
        fred_us,sc,signals,forex,crypto,
        sectors,us_stocks,tw_stocks,tw_flows,indices,stock_flows,rows_13f
    )

    log.info("="*60)
    log.info("  DONE")
    log.info("="*60)


if __name__=="__main__":
    run()
